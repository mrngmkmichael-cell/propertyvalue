"""Which districts has Google started showing that we do not prioritise?

Reads a Search Console performance export (the xlsx from Performance >
Export) and prints every real postcode district that earned impressions,
split into already-promoted and not-yet-promoted. The not-yet list is
printed ready to paste into GSC_EARNED_OUTCODES in app/main.py.

    .venv/Scripts/python.exe scripts/promote_from_gsc.py <export.xlsx>

This is the loop that grew the sitemap from 367 to 504 districts on
28 Aug 2026, made repeatable: the hand-picked list assumed big cities
mattered and the export showed the districts actually placing are
ordinary residential ones nobody else serves (the site's only click came
from /area/PE29 at position 4, which was not on the list). Run it on a
fresh export weekly; it changes nothing by itself.
"""
import collections
import os
import re
import sys
import textwrap

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

os.environ.setdefault("DATABASE_URL", "sqlite:///:memory:")
os.environ.setdefault("SESSION_SECRET", "promote-from-gsc")

import openpyxl  # noqa: E402

from app.main import AREA_GUIDE_SEED_OUTCODES, KNOWN_OUTCODES  # noqa: E402


def main() -> int:
    if len(sys.argv) < 2 or not os.path.exists(sys.argv[1]):
        print(__doc__)
        return 1
    wb = openpyxl.load_workbook(sys.argv[1], data_only=True)
    for sheet in ("Pages", "Queries"):
        if sheet not in wb.sheetnames:
            print(f"no {sheet!r} sheet: is this really a Performance export?")
            return 1

    found: collections.Counter = collections.Counter()
    for row in list(wb["Pages"].iter_rows(values_only=True))[1:]:
        if not row[0]:
            continue
        m = re.match(
            r"https://ukpropertyinsight\.co\.uk/(?:area/([A-Z0-9]+)(?:/|$)|schools/guide\?q=([A-Z0-9]+))",
            str(row[0]),
        )
        if m:
            found[m.group(1) or m.group(2)] += row[2] or 0
    for row in list(wb["Queries"].iter_rows(values_only=True))[1:]:
        if not row[0]:
            continue
        # A bare district typed into Google ("s20 sheffield", "b8").
        for token in re.findall(r"\b([a-z]{1,2}\d{1,2}[a-z]?)\b", str(row[0]).lower()):
            if token.upper() in KNOWN_OUTCODES:
                found[token.upper()] += row[2] or 0

    seed = set(AREA_GUIDE_SEED_OUTCODES)
    real = {oc: n for oc, n in found.items() if oc in KNOWN_OUTCODES}
    new = sorted(oc for oc in real if oc not in seed)
    already = sorted(oc for oc in real if oc in seed)

    print(f"{len(real)} districts earned impressions in this export")
    print(f"  already promoted : {len(already)}")
    print(f"  NEW              : {len(new)}\n")
    if not new:
        print("nothing to add: every earning district is already in the sitemap")
        return 0
    for oc in sorted(new, key=lambda o: -real[o]):
        print(f"  {oc:<7} {real[oc]:.0f} impressions")
    print("\npaste into GSC_EARNED_OUTCODES in app/main.py:\n")
    print(textwrap.fill(
        ", ".join(f'"{o}"' for o in new), width=74,
        initial_indent="    ", subsequent_indent="    ",
    ) + ",")
    print("\nthen run the tests and deploy; the sitemap and the comparison-page")
    print("tranche both grow from that one list.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
