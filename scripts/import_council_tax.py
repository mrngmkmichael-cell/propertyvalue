"""Import council tax levels for England, Scotland and Wales.

England: MHCLG "Council Tax levels set by local authorities in England"
(Table 10, Data_Billing, line 17: average Band D area charge including
all precepts), keyed by ONS code. Other bands derive from the statutory
English ninths at read time.

Scotland: Scottish Government "Council Tax by Band" workbook, which
publishes every band A-H per council directly (Scotland's own post-2017
ratios), keyed by council name as postcodes.io reports it.

Wales: Welsh Government "Council tax levels" release, Table 1 overall
average Band D per billing authority, keyed by name; bands A-I derive
from the Welsh statutory ninths at read time.

All three publish each March; update the URLs and re-run, then commit
the refreshed JSON.

Usage: python scripts/import_council_tax.py
"""
import json
import sys
import zipfile
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET

import httpx

ODS_URL = "https://assets.publishing.service.gov.uk/media/6a02eeeccd2e0e8b5b20b449/Table_10_2026-27.ods"
SCOTLAND_URL = ("https://www.gov.scot/binaries/content/documents/govscot/publications/statistics/2019/04/"
                "council-tax-datasets/documents/average-council-tax-per-dwelling/council-tax-by-band-2026-27/"
                "council-tax-by-band-2026-27/govscot%3Adocument/"
                "CTAS%2B2026%2B-%2BCouncil%2BTax%2BAssumptions%2B-%2BCouncil%2BTax%2Bby%2BBand%2B-%2B2026-27.xlsx")
WALES_URL = ("https://www.gov.wales/sites/default/files/statistics-and-research/2026-03/"
             "council-tax-levels-april-2026-march-2027-153478.xlsx")
YEAR_LABEL = "2026-27"
OUT = Path(__file__).resolve().parent.parent / "app" / "data" / "council_tax.json"

NS = {"table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
      "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0"}
T = "{urn:oasis:names:tc:opendocument:xmlns:table:1.0}"


def cells(row, cap=60):
    out = []
    for c in row.findall("table:table-cell", NS):
        rep = int(c.get(T + "number-columns-repeated", "1"))
        txt = " ".join("".join(p.itertext()) for p in c.findall("text:p", NS))
        out.extend([txt] * min(rep, cap))
        if len(out) > cap:
            break
    return out[:cap]


def main() -> None:
    print(f"Downloading {ODS_URL}")
    ods = httpx.get(ODS_URL, timeout=120, follow_redirects=True).content
    root = ET.fromstring(zipfile.ZipFile(BytesIO(ods)).read("content.xml"))
    tbl = next(t for t in root.findall(".//table:table", NS)
               if t.get(T + "name") == "Data_Billing")
    rows = [cells(r) for r in tbl.findall("table:table-row", NS)]
    header = rows[4]
    col = next(i for i, h in enumerate(header)
               if h.startswith("17. Average (Band D 2 adult equivalent) council tax for area"))
    print(f"Band D area-charge column: {col}")

    data = {}
    for r in rows[5:]:
        if len(r) <= col:
            continue
        ons, name, value = r[1].strip(), r[2].strip(), r[col].strip().replace(",", "")
        if not ons.startswith("E0") or not value:
            continue
        try:
            data[ons] = {"authority": name, "band_d": round(float(value), 2)}
        except ValueError:
            continue

    assert len(data) > 250, f"only {len(data)} authorities parsed - source layout changed?"
    sample = data.get("E07000223")
    assert sample and 1000 < sample["band_d"] < 4000, f"Adur sanity check failed: {sample}"

    scotland = scotland_bands()
    wales = wales_band_d()
    OUT.write_text(json.dumps({
        "year": YEAR_LABEL,
        "england": {"source": ODS_URL, "authorities": data},
        "scotland": {"source": SCOTLAND_URL, "authorities": scotland},
        "wales": {"source": WALES_URL, "authorities": wales},
    }, indent=1), encoding="utf-8")
    print(f"Wrote England {len(data)}, Scotland {len(scotland)}, Wales {len(wales)} to {OUT}")


def _norm(name: str) -> str:
    return " ".join(name.replace("&", "and").split()).lower()


def scotland_bands() -> dict:
    """All bands A-H per Scottish council, straight from the workbook -
    Scotland's post-2017 ratios are baked into the published figures."""
    import openpyxl
    print(f"Downloading {SCOTLAND_URL}")
    content = httpx.get(SCOTLAND_URL, timeout=120, follow_redirects=True,
                        headers={"user-agent": "Mozilla/5.0 ukpropertyinsight-import"}).content
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    start = next(i for i, r in enumerate(rows) if r[0] and "Ratio to Band D" in str(r[0])) + 1
    out = {}
    for r in rows[start:]:
        name = (r[0] or "").strip() if isinstance(r[0], str) else None
        if not name or "average" in name.lower():
            continue
        try:
            bands = {b: round(float(r[i + 1]), 2) for i, b in enumerate("ABCDEFGH")}
        except (TypeError, ValueError):
            continue
        out[_norm(name)] = {"authority": name, "bands": bands}
    assert len(out) == 32, f"expected 32 Scottish councils, got {len(out)}"
    return out


def wales_band_d() -> dict:
    """Overall average Band D per Welsh billing authority (Table 1)."""
    import openpyxl
    print(f"Downloading {WALES_URL}")
    content = httpx.get(WALES_URL, timeout=120, follow_redirects=True,
                        headers={"user-agent": "Mozilla/5.0 ukpropertyinsight-import"}).content
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb["Table1"]
    out = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        name = (r[0] or "").strip() if isinstance(r[0], str) else None
        if not name or "wales" in name.lower():
            continue
        try:
            band_d = round(float(r[1]), 2)
        except (TypeError, ValueError):
            continue
        out[_norm(name)] = {"authority": name, "band_d": band_d}
    assert len(out) == 22, f"expected 22 Welsh authorities, got {len(out)}"
    return out


if __name__ == "__main__":
    sys.exit(main())
