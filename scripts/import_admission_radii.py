"""One-time/periodic offline import of MODELLED school admission
radius estimates into the `school_admission_radii` table.

NOT run by the deployed app - run manually from a dev machine. Needs
`pdfplumber`, `openpyxl`, and `python-docx` installed locally (not
production dependencies - only used here, for extracting tables from
local authorities' own PDF/Excel/Word "last distance offered"
publications, same situation as pyproj in import_schools.py).

Unlike every other import script in this project, there is no single
source. Each English local authority publishes its own "how places
were offered" report in its own format (PDF table, prose, km, miles,
single or multi-year), at its own URL that typically changes every
year, with no central index. This script is a registry - one fetch
function per authority - meant to be extended authority by authority
over time, not filled in all at once. See SchoolAdmissionRadius's
docstring in app/models.py for what this data actually represents
(a modelled circle, not a real catchment boundary) and why.

To add another authority: write a fetch_<name>() function returning
[{"school_name": ..., "last_distance_miles": ...}, ...], then add it
to _AUTHORITIES below with the authority name (must exactly match
SchoolDetail.local_authority for that area) and academic year label.
"""
import csv
import difflib
import io
import os
import re
import sys

import docx
import httpx
import openpyxl
import pdfplumber

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dotenv import load_dotenv

load_dotenv()

from app.db import Base, _get_engine  # noqa: E402
from app.models import School, SchoolAdmissionRadius, SchoolDetail  # noqa: E402
from sqlalchemy import select  # noqa: E402
from sqlalchemy.orm import sessionmaker  # noqa: E402

HEADERS = {"User-Agent": "Mozilla/5.0"}
_DISTANCE_RE = re.compile(r"last distance[^\d]*([\d.]+)", re.IGNORECASE)

# Common school-name abbreviations that don't fuzzy-match well
# character-for-character against GIAS's spelled-out EstablishmentName
# (e.g. "Feltham Hill I&N" vs "Feltham Hill Infant and Nursery School").
_ABBREVIATIONS = {
    r"\bI&N\b": "Infant and Nursery",
    r"\bJ I & N\b": "Junior Infant and Nursery",
    r"\bJ & I\b": "Junior and Infant",
    r"\bI & N\b": "Infant and Nursery",
    r"\bJun\b": "Junior",
    r"\bInf\b": "Infant",
    r"\bJnr\b": "Junior",
    r"\bJMI\b": "Junior Mixed Infant",
    r"\bC of E\b": "Church of England",
    r"\bCEVCP\b": "Church of England Voluntary Controlled Primary",
    r"\bCE\b": "Church of England",
    r"\bVC\b": "Voluntary Controlled",
    r"\bVA\b": "Voluntary Aided",
    r"\bRC\b": "Roman Catholic",
    r"\bR\.C\.(?=\s|$)": "Roman Catholic",
    r"\bRd\b": "Road",
    r"\bCP\b": "Community Primary",
    r"\bC\.P\.(?=\s|$)": "Community Primary",
}

# Confidence floor for fuzzy name matching - below this, skip the row
# rather than guess. A wrong match here would draw the wrong school's
# admission circle on the wrong school, which is worse than drawing
# nothing.
_MATCH_CUTOFF = 0.72


def _normalize_school_name(name: str) -> str:
    for pattern, replacement in _ABBREVIATIONS.items():
        name = re.sub(pattern, replacement, name, flags=re.IGNORECASE)
    name = re.sub(r"\bschool\b", "", name, flags=re.IGNORECASE)
    name = re.sub(r"[^a-z0-9 ]", "", name.lower())
    return re.sub(r"\s+", " ", name).strip()


def fetch_hounslow() -> list[dict]:
    """London Borough of Hounslow's "How primary places were offered"
    PDF - republished each autumn at a new file ID, so find the
    current one via hounslow.gov.uk's own site search
    ("last distance offered") before re-running, and update the URL
    below if it 404s.
    """
    url = (
        "https://www.hounslow.gov.uk/downloads/file/11494/"
        "how-primary-places-were-offered-in-september-2025"
    )
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or len(row) < 6 or not row[0] or not row[5]:
                    continue
                match = _DISTANCE_RE.search(row[5].replace("\n", " "))
                if match:
                    records.append({"school_name": row[0].strip(), "last_distance_miles": float(match.group(1))})
    return records


_METRES_PER_MILE = 1609.34

# Matches a heading-style line - ALL CAPS, allowing digits/punctuation
# a school name could contain - used by the text-scan parsers below
# for councils that don't publish a clean table (school name as a
# standalone heading line, followed by free-form prose/mini-tables
# ending in a distance figure).
_HEADING_RE = re.compile(r"^[A-Z0-9][A-Z0-9\s&',.\-]{4,60}$")
_HEADING_SKIP_WORDS = ("OFFICIAL", "PAGE", "ADMISSIONS", "APRIL", "SCHOOL PLACES", "ALLOCATED", "OFFERED", "COUNCIL")
_METRES_OR_MILES_RE = re.compile(r"([\d,]+\.?\d*)\s*(metres|miles|km|m|mi)\b", re.IGNORECASE)


def _scan_headings_for_distance(full_text: str, unit_to_miles: float) -> list[dict]:
    """Generic parser for the "ALL CAPS SCHOOL NAME heading, then
    free-form text/mini-table ending in a distance figure, next
    heading" layout (first seen in Wandsworth's PDF, likely to recur -
    several other authorities format theirs the same way). Takes the
    LAST distance-shaped number before the next heading, since these
    documents put the final "furthest distance offered" figure at the
    end of each school's section."""
    lines = [ln.strip() for ln in full_text.split("\n") if ln.strip()]
    records = []
    current = None
    buffer: list[str] = []

    def _flush():
        if not current or not buffer:
            return
        text = " ".join(buffer)
        matches = list(_METRES_OR_MILES_RE.finditer(text))
        if matches:
            value = float(matches[-1].group(1).replace(",", ""))
            records.append({"school_name": current, "last_distance_miles": value / unit_to_miles})

    for line in lines:
        is_heading = _HEADING_RE.match(line) and not any(w in line for w in _HEADING_SKIP_WORDS)
        if is_heading:
            _flush()
            current = line
            buffer = []
        else:
            buffer.append(line)
    _flush()
    return records


def fetch_wandsworth() -> list[dict]:
    """London Borough of Wandsworth's "How places were offered at
    each Wandsworth school" PDF - republished each spring at a new
    URL each year (find the current one via wandsworth.gov.uk's own
    site search, "how places were allocated"). Distances published
    in metres; each school's name is a standalone heading line rather
    than a table column, so this uses the generic heading-scan parser.
    """
    url = "https://wandsworth.gov.uk/media/xonfunuu/how_places_were_allocated_for_primary_schools_2025.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    full_text = ""
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"
    return _scan_headings_for_distance(full_text, _METRES_PER_MILE)


def fetch_brent() -> list[dict]:
    """London Borough of Brent's "How places were offered - reception"
    PDF - republished each spring at a new file ID (find the current
    one via brent.gov.uk's own search). Table has one row per school
    with a multi-line cell listing every admission criterion (sibling,
    staff, distance, etc.) each with its own distance figure - this
    takes the largest (i.e. the true "furthest anyone got in" figure,
    not just the distance-criterion row) rather than guessing which
    line is which criterion from the flattened cell text.
    """
    url = "https://www.brent.gov.uk/media/16421130/how-places-were-offered-reception-2024.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or len(row) < 3 or not row[0] or not row[2]:
                    continue
                name = row[0].split("\n")[0].strip()
                distances = [float(d) for d in re.findall(r"[\d.]+", row[2])]
                if name and distances:
                    records.append({"school_name": name, "last_distance_miles": max(distances) / _METRES_PER_MILE})
    return records


def fetch_bristol() -> list[dict]:
    """Bristol City Council's "furthest distance table" - a multi-year
    time series (one column per year, 'D' where distance wasn't the
    deciding criterion that year) rather than a single latest-year
    figure. Takes the most recent column with a real value, per
    school, since a 'D' year doesn't mean no distance was recorded -
    it means the school wasn't oversubscribed enough for distance to
    matter that year, which isn't informative for this estimate.
    Republished periodically at bristol.gov.uk - re-check the file ID
    if this URL 404s.
    """
    url = "https://www.bristol.gov.uk/files/documents/3382-furthest-distance-table/file"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or not row[0] or row[0].strip().lower().startswith("name"):
                    continue
                name = row[0].strip()
                # Columns after the name are years, most recent last.
                for cell in reversed(row[1:]):
                    if cell and cell.strip().upper() != "D":
                        try:
                            records.append({"school_name": name, "last_distance_miles": float(cell.strip())})
                        except ValueError:
                            continue
                        break
    return records


def fetch_leeds() -> list[dict]:
    """Leeds City Council's primary school cut-off distances PDF -
    republished each spring at a new URL (find the current one via
    leeds.gov.uk's own search, "cut off distances"). Already in
    miles, one clean row per school - the most directly parseable
    format found so far. Rows reading "All applicants admitted" have
    no distance (not oversubscribed) and are correctly skipped.
    """
    url = "https://www.leeds.gov.uk/sites/default/files/2025-04/primary%20school%20cut%20off%20distances.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or len(row) < 3 or not row[1] or not row[2]:
                    continue
                match = re.match(r"[\d.]+", row[2].strip())
                if match:
                    name = row[1].replace("\n", " ").strip()
                    records.append({"school_name": name, "last_distance_miles": float(match.group(0))})
    return records


def fetch_kirklees() -> list[dict]:
    """Kirklees Council's "Reception by preference and criteria" PDF
    (Huddersfield/Dewsbury area, West Yorkshire) - republished each
    spring at a new file ID (find the current one via kirklees.gov.uk
    admissions pages). Cleanest format found so far: one row per
    school, last column is literally "Distance of the last on-time
    place allocated (metres)".
    """
    url = "https://www.kirklees.gov.uk/beta/admissions/pdf/reception-by-preference-and-criteria-25.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or len(row) < 2 or not row[0]:
                    continue
                distance_cell = (row[-1] or "").strip()
                try:
                    distance = float(distance_cell)
                except ValueError:
                    continue
                name = row[0].replace("\n", " ").strip()
                records.append({"school_name": name, "last_distance_miles": distance / _METRES_PER_MILE})
    return records


_EALING_DISTANCE_RE = re.compile(r"([\d.]+)\s*(of a mile|miles?)", re.IGNORECASE)
_EALING_EXCLUDE_PREFIXES = ("criteria", "number", "places allocated", "no supplementary")


def fetch_ealing() -> list[dict]:
    """London Borough of Ealing's "Primary school on-time offers"
    PDF - republished each spring at a new file ID (find the current
    one via ealing.gov.uk's own search). Already in miles. Most
    schools are one clean row; a handful of faith schools break their
    admissions down by sub-criterion across several rows - this
    tracks the most recent school-name row seen and attaches the
    distance figure from whichever row it actually appears on
    (usually the last oversubscribed criterion), rather than assuming
    a fixed column position that only holds for the simple schools.
    """
    url = "https://www.ealing.gov.uk/download/downloads/id/18843/primary_school_on-time_offers_2025.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            current_name = None
            for row in table:
                cells = [c for c in row if c]
                if not cells:
                    continue
                first = cells[0].strip()
                if not first.lower().startswith(_EALING_EXCLUDE_PREFIXES) and "primary school" in first.lower():
                    current_name = first
                match = _EALING_DISTANCE_RE.search(" ".join(cells))
                if match and current_name:
                    records.append({"school_name": current_name, "last_distance_miles": float(match.group(1))})
                    current_name = None  # one distance per school - stop after the first hit
    return records


def fetch_hackney() -> list[dict]:
    """London Borough of Hackney's "How Places Were Offered - Reception"
    PDF - republished each spring at a new URL each year (find the
    current one via hackney.gov.uk/education's own search). Cleanest
    format found yet: one row per school, second-to-last column is
    literally "Max Distance (last child offered in miles)", already
    in miles.
    """
    url = "https://education.hackney.gov.uk/sites/default/files/document/How%20Places%20Were%20Offered%20-%20Reception%202025.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or len(row) < 12 or not row[0] or not row[-2]:
                    continue
                try:
                    distance = float(row[-2].strip())
                except ValueError:
                    continue
                records.append({"school_name": row[0].strip(), "last_distance_miles": distance})
    return records


def fetch_solihull() -> list[dict]:
    """Solihull Council's "How Reception Places Were Offered" PDF - a
    3-year time series (like Bristol's), republished each spring at a
    new URL (find the current one via solihull.gov.uk's own search).
    Takes the most recent year's figure ("All offered"/N/A means not
    oversubscribed that year, so falls back to the next most recent
    year with a real value).
    """
    url = "https://www.solihull.gov.uk/sites/default/files/2025-04/How-Reception-Places-Were-Offered-23-24-25.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or not row[0] or row[0].strip().lower() in ("", "school"):
                    continue
                name = row[0].strip()
                # Distance columns are at indices 2, 4, 6 (most recent first).
                for idx in (2, 4, 6):
                    if idx >= len(row) or not row[idx]:
                        continue
                    try:
                        records.append({"school_name": name, "last_distance_miles": float(row[idx].strip())})
                        break
                    except ValueError:
                        continue
    return records


def fetch_harrow() -> list[dict]:
    """London Borough of Harrow's "How places were allocated at
    Harrow Schools" PDF - republished each spring at a new file ID
    (find the current one via harrow.gov.uk's own search). One clean
    row per school; "FURTHEST DISTANCE OFFERED IN MILES..." is the
    third-from-last column, already in miles.
    """
    url = "https://www.harrow.gov.uk/downloads/file/33045/Reception_2025___How_places_were_allocated_at_Harrow_Schools.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or len(row) < 3 or not row[0] or not row[-3]:
                    continue
                try:
                    distance = float(row[-3].strip())
                except ValueError:
                    continue
                records.append({"school_name": row[0].replace("\n", " ").strip(), "last_distance_miles": distance})
    return records


def fetch_gloucestershire() -> list[dict]:
    """Gloucestershire County Council's "Primary allocation day
    statistics" PDF - republished each spring at a new file ID (find
    the current one via gloucestershire.gov.uk's own search). One
    clean row per school; "Furthest distance allocated (miles)" is
    column index 5, already in miles. A county council rather than a
    single city/borough - broader geographic spread than the mostly
    urban authorities covered so far.
    """
    url = "https://www.gloucestershire.gov.uk/media/ctgfusdi/primary-allocation-day-statistics-2025-1.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or len(row) < 6 or not row[1] or not row[5]:
                    continue
                try:
                    distance = float(row[5].strip())
                except ValueError:
                    continue
                records.append({"school_name": row[1].replace("\n", " ").strip(), "last_distance_miles": distance})
    return records


_BIRMINGHAM_DISTANCE_RE = re.compile(r"([\d,]+\.?\d*)\s*metres", re.IGNORECASE)


def fetch_birmingham() -> list[dict]:
    """Birmingham City Council's "Primary offers" PDF - republished
    each year at a new file ID (find the current one via
    birmingham.gov.uk's own search, "breakdown of primary school
    offers"). One row per school; the last column is "Cut Off
    Distance YYYY (where applicable)" - sometimes a plain "676
    metres", sometimes "Faith (4166 metres)" (still a real distance,
    just alongside the admitting criterion) - a regex extracts the
    number regardless of what text surrounds it, and rows reading
    "All Applicants" or similar with no metres figure are correctly
    skipped. England's second-largest city by population.
    """
    url = "https://www.birmingham.gov.uk/download/downloads/id/29456/primary_offers_2024.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or not row[0] or not row[-1]:
                    continue
                match = _BIRMINGHAM_DISTANCE_RE.search(row[-1].replace(",", ""))
                if match:
                    name = row[0].replace("\n", " ").strip()
                    records.append({"school_name": name, "last_distance_miles": float(match.group(1)) / _METRES_PER_MILE})
    return records


_NEWCASTLE_RECEPTION_ROW_RE = re.compile(
    r'<p id="([^"]+)"><strong>[^<]+</strong></p>.*?'
    r'Farthest distance from school</td><td[^>]*>([^<]*)</td>',
    re.DOTALL,
)
_NEWCASTLE_TRANSFER_SCHOOL_RE = re.compile(r"<h3>([^<]+)</h3>(.*?)(?=<h3>|\Z)", re.DOTALL)
_NEWCASTLE_DISTANCE_MILES_RE = re.compile(r"([\d.]+)\s*miles?", re.IGNORECASE)


def fetch_newcastle() -> list[dict]:
    """Newcastle City Council publishes its "how places were allocated"
    results as plain HTML pages (not PDFs) - one for Reception
    (primary) and one for Transfer (secondary), both republished each
    year at a new URL suffix (find the current ones via
    newcastle.gov.uk's own admissions pages). Reception page is one
    big table with a "Farthest distance from school" row per school,
    already in miles; Transfer page uses an "<h3>School Name</h3>"
    heading per school followed by a "Last distance offered in
    Category N: X.XXX miles" paragraph. Schools reading "School not
    filled on distance criteria" have no distance and are correctly
    skipped.
    """
    records = []

    reception_url = "https://newcastle.gov.uk/services/how-reception-places-were-allocated-2026"
    print(f"  Downloading {reception_url}")
    resp = httpx.get(reception_url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    for name, value in _NEWCASTLE_RECEPTION_ROW_RE.findall(resp.text):
        match = _NEWCASTLE_DISTANCE_MILES_RE.search(value)
        if match:
            records.append({"school_name": name.strip(), "last_distance_miles": float(match.group(1))})

    transfer_url = (
        "https://www.newcastle.gov.uk/services/schools-learning-and-childcare/"
        "apply-school-place/information-about-how-reception-and-6"
    )
    print(f"  Downloading {transfer_url}")
    resp = httpx.get(transfer_url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    for name, section in _NEWCASTLE_TRANSFER_SCHOOL_RE.findall(resp.text):
        match = _NEWCASTLE_DISTANCE_MILES_RE.search(section)
        if match:
            records.append({"school_name": name.strip(), "last_distance_miles": float(match.group(1))})

    return records


_SURREY_SCHOOL_RE = re.compile(
    r"^([A-Za-z][^\n]{2,90})\nDfE No: [\d/]+\n(.*?)(?=\n[A-Za-z][^\n]{2,90}\nDfE No: |\Z)",
    re.MULTILINE | re.DOTALL,
)
_SURREY_DISTANCE_RE = re.compile(r"Distance[^=\n]{0,25}=\s*([\d.]+)\s*km")

# One PDF per district (plus one for Junior schools, one for
# Secondary) - Surrey is a county council (like Gloucestershire) so
# GIAS's local_authority for all of these is simply "Surrey", but the
# council itself only publishes allocation figures broken down by
# district, hence the long URL list.
_SURREY_URLS = [
    "https://www.surreycc.gov.uk/__data/assets/pdf_file/0010/521389/FINAL-Elmbridge-Primary-allocation-figures-September-2026-V1.pdf",
    "https://www.surreycc.gov.uk/__data/assets/pdf_file/0020/521390/FINAL-Epsom-and-Ewell-Primary-allocation-figures-September-2026-V1.pdf",
    "https://www.surreycc.gov.uk/__data/assets/pdf_file/0003/521391/FINAL-Guildford-Primary-allocation-figures-September-2026-V1.pdf",
    "https://www.surreycc.gov.uk/__data/assets/pdf_file/0004/521392/FINAL-Mole-Valley-Primary-allocation-figures-September-2026-V1.pdf",
    "https://www.surreycc.gov.uk/__data/assets/pdf_file/0005/521393/FINAL-Reigate-and-Banstead-Primary-allocation-figures-September-2026-V1.pdf",
    "https://www.surreycc.gov.uk/__data/assets/pdf_file/0006/521394/FINAL-Runnymede-Primary-allocation-figures-September-2026-V1.pdf",
    "https://www.surreycc.gov.uk/__data/assets/pdf_file/0007/521395/FINAL-Spelthorne-Primary-allocation-figures-September-2026-V1.pdf",
    "https://www.surreycc.gov.uk/__data/assets/pdf_file/0008/521396/FINAL-Surrey-Heath-Primary-allocation-figures-September-2026-V1.pdf",
    "https://www.surreycc.gov.uk/__data/assets/pdf_file/0009/521397/FINAL-Tandridge-Primary-allocation-figures-September-2026-V1.pdf",
    "https://www.surreycc.gov.uk/__data/assets/pdf_file/0010/521398/FINAL-Waverley-Primary-allocation-figures-September-2026-V1.pdf",
    "https://www.surreycc.gov.uk/__data/assets/pdf_file/0011/521399/FINAL-Woking-Primary-allocation-figures-September-2026-V1.pdf",
    "https://www.surreycc.gov.uk/__data/assets/pdf_file/0009/521388/FINAL-Junior-allocation-figures-September-2026-V1.pdf",
    "https://www.surreycc.gov.uk/__data/assets/pdf_file/0003/516684/FINAL-Secondary-allocation-figures-September-2026-V2.pdf",
]


def _parse_surrey_pdf(full_text: str) -> list[dict]:
    records = []
    for name, section in _SURREY_SCHOOL_RE.findall(full_text):
        # "Distance to Nodal Point" is a distance to a fixed
        # reference point used as a tie-breaker on some split
        # catchments, not a home-to-school distance - excluded so it
        # doesn't get mistaken for the admission radius.
        distances = [
            float(m.group(1))
            for m in _SURREY_DISTANCE_RE.finditer(section)
            if "nodal" not in m.group(0).lower()
        ]
        if distances:
            records.append({"school_name": name.strip(), "last_distance_miles": max(distances) / 1.60934})
    return records


def fetch_surrey() -> list[dict]:
    """Surrey County Council's "Allocation of places" PDFs - one per
    district (Elmbridge, Epsom and Ewell, Guildford, Mole Valley,
    Reigate and Banstead, Runnymede, Spelthorne, Surrey Heath,
    Tandridge, Waverley, Woking) plus separate Junior and Secondary
    documents, republished each year at new file IDs under the same
    /__data/assets/pdf_file/ path pattern (find the current ones via
    surreycc.gov.uk's own "arrangements-and-outcomes/previous-years"
    page). Format is prose, not a table: each school is a heading line
    followed by "DfE No: .../....", then free text ending in
    "Distance = X.XXXkm" (already scoped to the last-filled criterion
    per the document's own key) - already in km, converted to miles.
    """
    records = []
    for url in _SURREY_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        full_text = ""
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"
        records.extend(_parse_surrey_pdf(full_text))
    return records


_HERTS_NAME_RE = re.compile(r"^(.+?)\s+Total Applications:\s*\d+", re.MULTILINE)
_HERTS_DISTANCE_RE = re.compile(r"distance of most distant child admitted\s*([\d,]+\.?\d*)\s*m\b")

_HERTS_URLS = [
    "https://www.hertfordshire.gov.uk/doc/sch/adm/stats/primary-allocation-summary-reports-26-27.pdf",
    "https://www.hertfordshire.gov.uk/doc/sch/adm/stats/junior-and-middle-school-allocation-summary-reports-26-27.pdf",
    "https://www.hertfordshire.gov.uk/media-library/documents/schools-and-education/admissions/"
    "previous-years-stats/25-26/allocation-summary-reports.pdf",
]


def fetch_hertfordshire() -> list[dict]:
    """Hertfordshire County Council's "School Allocation Summary
    Report" PDFs - one for Primary, one for Junior/Middle, one for
    Secondary/Upper (the secondary one is still on last year's 25-26
    URL since 26-27's hadn't been published yet at time of writing -
    check hertfordshire.gov.uk's "previous years' statistics" page for
    a newer one). One page per school (a large multi-hundred-page
    document, unlike every other authority so far), always titled
    "<School Name> Total Applications: N", with the actual metric
    being "Home to school distance of most distant child admitted X m"
    - can appear more than once per school (e.g. separate "Nearest
    School"/"Not Nearest School" or "In Priority Area" rows), so this
    takes the largest, consistent with the same choice made for other
    authorities with multiple distance-shaped rows per school (e.g.
    Brent, Surrey). Already in metres.
    """
    records = []
    for url in _HERTS_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                name_match = _HERTS_NAME_RE.search(text)
                if not name_match:
                    continue
                distances = [float(d.replace(",", "")) for d in _HERTS_DISTANCE_RE.findall(text)]
                if distances:
                    records.append({
                        "school_name": name_match.group(1).strip(),
                        "last_distance_miles": max(distances) / _METRES_PER_MILE,
                    })
    return records


_OXFORDSHIRE_DISTANCE_RE = re.compile(r"([\d.]+)\s*miles?", re.IGNORECASE)


def fetch_oxfordshire() -> list[dict]:
    """Oxfordshire County Council publishes its "last offer" data as a
    plain CSV (not PDF!) - "Last place offered at schools that had
    refusals", republished each allocation round at a new file name
    (find the current one via oxfordshire.gov.uk's own
    "allocation-reports-and-vacancies/primary-allocation" page - the
    "P<year>-NOD<n>-LastOffer.csv" naming pattern is likely to recur).
    Establishment name has a trailing "(DfE No)" suffix stripped
    before matching; distance is already in miles. Cleanest source
    format found across every authority in this registry so far.
    """
    url = "https://www.oxfordshire.gov.uk/sites/default/files/file/place-allocations/P26-NOD1-LastOffer.csv"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    text = resp.content.decode("utf-8-sig", errors="replace")

    records = []
    for line in text.splitlines()[2:]:
        if not line.strip():
            continue
        cells = next(csv.reader([line]))
        if len(cells) < 6 or not cells[0]:
            continue
        name = re.sub(r"\s*\(\d+/\d+\)\s*$", "", cells[0]).strip()
        match = _OXFORDSHIRE_DISTANCE_RE.search(cells[5])
        if match:
            records.append({"school_name": name, "last_distance_miles": float(match.group(1))})
    return records


_BUCKS_MILES_RE = re.compile(r"([\d.]+)\s*miles?", re.IGNORECASE)

# (label, URL, "table" = row-per-school with its own distance column,
#  "prose" = row-per-school with a free-text description ending in a
#  distance figure - the Reception/Junior workbooks use the former,
#  the Secondary one the latter)
_BUCKS_FILES = [
    ("table", "https://www.buckinghamshire.gov.uk/documents/41709/"
              "How_Places_Were_Allocated_Into_Reception_September_2026_as_at_20_May_2026.xlsx"),
    ("table", "https://www.buckinghamshire.gov.uk/documents/41714/"
              "How_Places_Were_Allocated_Into_Junior_School_September_2026_as_at_20_May_2026.xlsx"),
    ("prose", "https://www.buckinghamshire.gov.uk/documents/41705/"
              "Allocation_Profile_2026_-_Secondary_Second_Round_-_20_May.xlsx"),
]


def fetch_buckinghamshire() -> list[dict]:
    """Buckinghamshire Council publishes its "How places were
    allocated" results as Excel workbooks (not PDFs), republished each
    year at a new document ID (find the current ones via
    buckinghamshire.gov.uk's "school-place-allocation-statistics"
    page). Reception and Junior workbooks are a clean table with a
    "Distance of last allocated child" column already in miles.
    Secondary is a coarser "school name, free-text description of how
    places were allocated" layout (grammar/upper/all-ability schools
    grouped under section-header rows with no distance, which are
    naturally skipped since they don't match the distance regex) -
    text ends with "...to X.XXX miles." when the school was
    oversubscribed enough for distance to matter.
    """
    records = []
    for kind, url in _BUCKS_FILES:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0] or not isinstance(row[0], str):
                continue
            name = row[0].strip()
            if kind == "table":
                cell = row[2] if len(row) > 2 else None
            else:
                cell = row[1] if len(row) > 1 else None
            if not cell or not isinstance(cell, str):
                continue
            match = _BUCKS_MILES_RE.search(cell)
            if match:
                records.append({"school_name": name, "last_distance_miles": float(match.group(1))})
    return records


_CAMBS_DISTANCE_RE = re.compile(r"([\d.]+)")

# Only the "round 1" Reception and Junior PDFs - the cleanest of
# Cambridgeshire's several allocation-round documents (one table,
# "School Name, PAN, Offered, Criterion, Distance (miles)"). Round 2,
# Middle, and Year 7/9 documents use inconsistent/wider layouts not
# worth the extra parsing complexity for the schools they'd add.
_CAMBS_URLS = [
    "https://cambridgeshire.gov.uk/asset-library/Reception-allocation-sheet-2026-round-1.pdf",
    "https://cambridgeshire.gov.uk/asset-library/Junior-data-sheet-2026-round-1.pdf",
]


def fetch_cambridgeshire() -> list[dict]:
    """Cambridgeshire County Council's Reception and Junior "round 1"
    allocation PDFs - republished each year at a new URL under
    /asset-library/ (find the current ones via cambridgeshire.gov.uk's
    "school-allocation-information" page). Clean table: School Name,
    PAN, Offered, Criterion Allocated to, Distance (miles) - already
    in miles, last column.
    """
    records = []
    for url in _CAMBS_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table:
                    continue
                for row in table:
                    if not row or len(row) < 5 or not row[0] or not row[4]:
                        continue
                    match = _CAMBS_DISTANCE_RE.search(row[4])
                    if match:
                        records.append({"school_name": row[0].strip(), "last_distance_miles": float(match.group(1))})
    return records


def _wokingham_distance_column(header: list) -> int | None:
    """Wokingham's PDF renders rotated column headers as
    line-reversed text (pdfplumber artifact of a 90-degree-rotated
    header cell) - e.g. "Distance of child allocated last" comes out
    as "tsal detacolla dlihc fo ecnatsiD" split across lines. Reversing
    each line back locates the column - except in the widest
    (17-column) table, where merged header cells throw off the
    header/data column alignment by one, so that layout's distance
    index (9, confirmed by inspecting its data rows directly) is
    hardcoded instead."""
    if len(header or []) == 17:
        return 9
    for i, cell in enumerate(header or []):
        if not cell:
            continue
        for line in cell.split("\n"):
            if "distance" in line[::-1].lower():
                return i
    return None


def fetch_wokingham() -> list[dict]:
    """Wokingham Borough Council's "Allocation breakdown for admission
    to Primary School" PDF - republished each spring at a new URL
    under /sites/wokingham/files/YYYY-MM/ (find the current one via
    wokingham.gov.uk's "key-dates-and-statistics" page). No secondary
    equivalent is published. Two different table layouts appear across
    pages (a wide one for schools using the council's own coordinated
    criteria, a narrower one for schools that set their own admission
    criteria) - the distance column index differs between them, so
    _wokingham_distance_column() locates it per-table rather than
    assuming a fixed position. Distances over 20 miles are dropped as
    an extraction glitch (this table has cases where a missing cell
    shifts an unrelated large number, like an application count, into
    the distance column) - Wokingham is a small borough where a
    genuine cross-borough admission would never plausibly reach that
    far, unlike Surrey's 47km faith-school outlier which was a
    legitimate published figure in a clean, unambiguous column.
    """
    url = "https://www.wokingham.gov.uk/sites/wokingham/files/2026-04/Primary%20school%20allocation%20statistics%202026.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table or len(table) < 2:
                continue
            dist_idx = _wokingham_distance_column(table[0])
            if dist_idx is None:
                continue
            for row in table[1:]:
                if not row or not row[0] or dist_idx >= len(row) or not row[dist_idx]:
                    continue
                name = re.sub(r"\*+$", "", row[0]).strip()
                try:
                    distance = float(row[dist_idx].strip())
                except ValueError:
                    continue
                if 0 < distance <= 20:
                    records.append({"school_name": name, "last_distance_miles": distance})
    return records


_COVENTRY_ROW_RE = re.compile(r'<a[^>]*>([^<]+?)</a>.*?</tr>', re.DOTALL)
_COVENTRY_DISTANCE_RE = re.compile(r"distance of\s*([\d.]+)\s*(of a mile|miles?)", re.IGNORECASE)

_COVENTRY_URLS = [
    "https://www.coventry.gov.uk/school-admissions/primary-school-admissions/4",
    "https://coventry.gov.uk/school-admissions/secondary-school-admissions/7",
]


def fetch_coventry() -> list[dict]:
    """Coventry City Council publishes its allocation results as plain
    HTML tables (not PDFs) - "How primary school places were
    allocated" and "Allocation of Coventry secondary school places",
    republished each year at the same URL (just the year in the page
    text changes, so these URLs are likely stable). One row per
    school: name is a link to its admissions policy, last column is
    prose like "...some offered up to a distance of 1.83 miles" or
    "...0.408 of a mile" - both phrasings handled. Schools that
    weren't oversubscribed have no distance mentioned and are
    correctly skipped.
    """
    records = []
    for url in _COVENTRY_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        for name, row in [(m.group(1), m.group(0)) for m in _COVENTRY_ROW_RE.finditer(resp.text)]:
            match = _COVENTRY_DISTANCE_RE.search(row.replace("&nbsp;", " "))
            if match:
                records.append({"school_name": name.strip(), "last_distance_miles": float(match.group(1))})
    return records


_MK_DISTANCE_RE = re.compile(r"distance\s*(?:to the [\w\s]+ campus)?\s*of\s*([\d.]+)\s*miles?", re.IGNORECASE)

_MK_URLS = [
    "https://www.milton-keynes.gov.uk/sites/default/files/2025-04/Allocation%20Profile%20Starting%20School%2016%20April%202025.pdf",
    "https://www.milton-keynes.gov.uk/sites/default/files/2025-02/Allocation%20Profile%203%20March%202025.pdf",
]


def fetch_milton_keynes() -> list[dict]:
    """Milton Keynes City Council's "Allocation Profile" PDFs - one
    for Starting/Primary School, one for Year 7/Secondary, republished
    each year at a new URL under /sites/default/files/YYYY-MM/ (find
    the current ones via milton-keynes.gov.uk's admissions pages).
    One clean row per school: name, then prose ending "...to a
    distance of X.XXX miles" (occasionally "...to the <campus name>
    campus" for split-site schools, handled by the regex's optional
    middle clause), vacancy flag.
    """
    records = []
    for url in _MK_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table:
                    continue
                for row in table[1:]:
                    if not row or not row[0] or not row[1]:
                        continue
                    match = _MK_DISTANCE_RE.search(row[1].replace("\n", " "))
                    if match:
                        name = row[0].replace("\n", " ").strip()
                        records.append({"school_name": name, "last_distance_miles": float(match.group(1))})
    return records


_RBWM_DISTANCE_RE = re.compile(r"Furthest distance met:\s*([\d.]+)\s*miles")

_RBWM_URLS = [
    "https://5f2fe3253cd1dfa0d089-bf8b2cdb6a1dc2999fecbc372702016c.ssl.cf3.rackcdn.com/"
    "uploads/ckeditor/attachments/17730/Allocation_information_for_RBWM_primary_schools_September_2025_V1.pdf",
    "https://5f2fe3253cd1dfa0d089-bf8b2cdb6a1dc2999fecbc372702016c.ssl.cf3.rackcdn.com/"
    "uploads/ckeditor/attachments/17544/FINAL_NOD_afc_middle_secondary_and_upper_school_allocation_information_2025_v2.pdf",
]


def fetch_windsor_and_maidenhead() -> list[dict]:
    """Royal Borough of Windsor and Maidenhead's "Allocation
    information" PDFs (Primary + Middle/Secondary/Upper) - republished
    each year at a new attachment ID on a CDN host (find the current
    ones via rbwm.gov.uk's school admissions pages). Prose layout, one
    block per school: "<School Name>", "Type: ... Number of places
    offered: N", "DfE Ref: .../... Number of divert offers: N",
    "Furthest distance met: X.XXX miles ...". The school name is
    always exactly two lines above the "DfE Ref:" line, which is a
    more reliable anchor than scanning forward from the name (some
    names get split across a stray hyperlink line in extraction).
    """
    records = []
    for url in _RBWM_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        full_text = ""
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"
        lines = full_text.split("\n")
        for i, line in enumerate(lines):
            if not line.strip().startswith("DfE Ref:") or i < 2 or not lines[i - 1].strip().startswith("Type:"):
                continue
            name = lines[i - 2].strip()
            for j in range(i, min(i + 3, len(lines))):
                match = _RBWM_DISTANCE_RE.search(lines[j])
                if match:
                    records.append({"school_name": name, "last_distance_miles": float(match.group(1))})
                    break
    return records


_STOCKPORT_URLS = [
    "https://assets.ctfassets.net/ii3xdrqc6nfw/5wDH7mjJwVyUmBxuKNKGyW/68b5a6974d1c750554e2352e180ed271/"
    "Allocation_of_places_at_primary_schools_in_Stockport_for_September_2026.pdf",
    "https://assets.ctfassets.net/ii3xdrqc6nfw/4EAb1hwybiJ3vKTqKigTMQ/87e71c0afd97857f5f6125fd0f95503b/"
    "Allocation_of_places_at_secondary_schools_in_Stockport_for_September_2026.pdf",
]
_STOCKPORT_DECIMAL_RE = re.compile(r"^\d+\.\d+$")
_STOCKPORT_INT_RE = re.compile(r"^\d+$")


def fetch_stockport() -> list[dict]:
    """Stockport Council's "Allocation of places" PDFs (Primary +
    Secondary) - republished each year at a new content-hash URL under
    assets.ctfassets.net (find the current ones via
    stockport.gov.uk/documents/school-allocation). Extremely wide,
    sparse tables (40-58 columns, mostly empty, with rotated/wrapped
    header text split across several rows) make a fixed column index
    unreliable - instead, per row, the school name is the first
    non-empty cell that isn't a plain integer, and the distance is the
    one cell matching a decimal number (every other numeric column in
    these tables is a plain integer count, so this is unambiguous).
    Header/label rows naturally have no decimal cell and are skipped.
    """
    records = []
    for url in _STOCKPORT_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table:
                    continue
                for row in table:
                    if not row:
                        continue
                    cells = [c.strip() if c else None for c in row]
                    non_empty = [c for c in cells if c]
                    name = None
                    distance = None
                    for c in non_empty:
                        if _STOCKPORT_DECIMAL_RE.match(c):
                            distance = float(c)
                        elif name is None and not _STOCKPORT_INT_RE.match(c) and c.lower() != "miles":
                            name = c
                    if name and distance:
                        records.append({"school_name": name, "last_distance_miles": distance})
    return records


_PORTSMOUTH_ROW_RE = re.compile(r"<tr>\s*<td[^>]*>([^<]+)</td>\s*<td[^>]*>(.*?)</td>\s*</tr>", re.DOTALL)
_PORTSMOUTH_DISTANCE_RE = re.compile(r"lived\s*([\d.]+)\s*miles away", re.IGNORECASE)

_PORTSMOUTH_URLS = [
    "https://www.portsmouth.gov.uk/services/schools-and-learning/schools/admissions/"
    "waiting-list-and-appeals-for-a-school-place/starting-school-important-information/",
    "https://www.portsmouth.gov.uk/services/schools-learning-and-childcare/schools/admissions/"
    "waiting-list-and-appeals-for-a-school-place/important-information-for-secondary-transfer-applicants/",
]


def fetch_portsmouth() -> list[dict]:
    """Portsmouth City Council publishes an "Oversubscribed schools"
    HTML table (not a PDF) on its own admissions pages - one for
    starting/primary school, one for secondary transfer, both
    seemingly stable URLs updated in place each year (though note the
    two use slightly different URL path segments -
    "schools-and-learning" vs "schools-learning-and-childcare" -
    re-check both if either 404s). Each row is "<school name>, prose
    ending '...lived X.XXX miles away from the school.'" - schools
    filled without reaching the distance criterion aren't listed at
    all, so every row here already had a distance decide the last
    place.
    """
    records = []
    for url in _PORTSMOUTH_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        for name, cell in _PORTSMOUTH_ROW_RE.findall(resp.text):
            match = _PORTSMOUTH_DISTANCE_RE.search(cell)
            if match:
                records.append({"school_name": name.strip(), "last_distance_miles": float(match.group(1))})
    return records


_PETERBOROUGH_URLS = [
    "https://peterborough.gov.uk/asset-library/reception-allocations-sheet-2026-website.pdf",
    "https://peterborough.gov.uk/asset-library/junior-allocations-sheet-2026-website.pdf",
    "https://peterborough.gov.uk/asset-library/1st-round-allocation-stats-02.03.26-website.pdf",
]


def fetch_peterborough() -> list[dict]:
    """Peterborough City Council's Reception, Junior, and Year 7
    "1st round" allocation PDFs - republished each year at new URLs
    under /asset-library/ (find the current ones via
    peterborough.gov.uk's "school-allocation-information" page).
    Reception/Junior share one table layout (last column "Distance
    (miles)"); Year 7 uses a wider layout with the distance column at
    a different position - rather than hardcode two different column
    indices, this reuses the same decimal-cell-detection approach as
    Stockport (every other numeric column across all three files is a
    plain integer count or "N/A", so the one cell matching a decimal
    number is unambiguous). One row (John Clare Primary School) reads
    "206.404" - a genuine typo in the council's own PDF, not an
    extraction glitch (Peterborough is a small unitary; no primary
    admission is 200 miles), so implausible values (>30 miles) are
    dropped rather than propagated as a real "circle" onto the map.
    """
    records = []
    for url in _PETERBOROUGH_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table:
                    continue
                for row in table:
                    if not row or not row[0] or not isinstance(row[0], str):
                        continue
                    name = row[0].replace("\n", " ").strip()
                    distance = None
                    for cell in row[1:]:
                        if cell and _STOCKPORT_DECIMAL_RE.match(cell.strip()):
                            distance = float(cell.strip())
                            break
                    if distance is not None and distance <= 30:
                        records.append({"school_name": name, "last_distance_miles": distance})
    return records


_NORTH_SOMERSET_INDEX_URL = (
    "https://n-somerset.gov.uk/my-services/schools-learning/school-admissions/"
    "oversubscribed-schools/primary-allocations-2024-25"
)
_NORTH_SOMERSET_PDF_RE = re.compile(r'href="([^"]+\.pdf)"')
_NORTH_SOMERSET_NAME_RE = re.compile(r"Allocation sheet for\s+(.*?)\s*\d{4}/\d{2}", re.DOTALL)
_NORTH_SOMERSET_DISTANCE_RE = re.compile(
    r"distance between home and school for the last child offered a place was\s*([\d.]+)\s*miles", re.IGNORECASE
)


def fetch_north_somerset() -> list[dict]:
    """North Somerset Council publishes one individual "Allocation
    sheet" PDF per oversubscribed primary school (not one bulk
    document) - this crawls the year's index page for the current
    list of PDF links rather than hardcoding them individually, since
    which schools were oversubscribed (and thus which PDFs exist)
    changes every year (find the current index page via
    n-somerset.gov.uk's "oversubscribed-schools/previous-primary-
    allocations" page - the "primary-allocations-YYYY-YY" URL pattern
    is likely to recur). Only ~13 schools a year (a small unitary), but
    a clean, unambiguous prose format: "The distance between home and
    school for the last child offered a place was X.XXX miles, as
    measured in a direct line."
    """
    print(f"  Downloading index {_NORTH_SOMERSET_INDEX_URL}")
    resp = httpx.get(_NORTH_SOMERSET_INDEX_URL, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    pdf_urls = [u for u in _NORTH_SOMERSET_PDF_RE.findall(resp.text) if "allocation" in u.lower()]

    records = []
    for url in pdf_urls:
        print(f"  Downloading {url}")
        pdf_resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        pdf_resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(pdf_resp.content)) as pdf:
            full_text = (pdf.pages[0].extract_text() or "")
        name_match = _NORTH_SOMERSET_NAME_RE.search(full_text)
        distance_match = _NORTH_SOMERSET_DISTANCE_RE.search(full_text)
        if name_match and distance_match:
            name = name_match.group(1).replace("\n", " ").strip()
            records.append({"school_name": name, "last_distance_miles": float(distance_match.group(1))})
    return records


_SOMERSET_URLS = [
    "https://somersetcc.sharepoint.com/:b:/s/SCCPublic/"
    "IQAmnq1_wHcJRayL6PFb9cfLAYPlkUMlIjVAMyhRFmP_Fng?e=sFNh7V&download=1",
    "https://somersetcc.sharepoint.com/:b:/s/SCCPublic/"
    "IQAs4cxF933FT6CTjdVlHM51AZxGnuma-esY5uYFvCNptJA?e=Ls4sqa&download=1",
]
_SOMERSET_SCHOOL_RE = re.compile(r"Allocation Summary for\s+(.+?)\n")
_SOMERSET_DISTANCE_LINE_RE = re.compile(r"(?:^|\n)\d+\s+\S.*?\s(\d+)\s+([\d.]+)\s*(?=\n|$)")


def fetch_somerset() -> list[dict]:
    """Somerset County Council's "Allocation Summaries" PDFs (First
    Admissions/Primary + Secondary), hosted on SharePoint share links
    rather than the council's own domain - the share URL only returns
    a raw PDF with "&download=1" appended (otherwise it 200s with an
    HTML viewer page instead of the file), republished each year at a
    new share link (find the current ones via somerset.gov.uk's
    "school-place-allocation-summaries" page, inside the year's
    accordion section). One page per school: "Allocation Summary for
    <Name>", then a "Criterion | Number of places offered | Max
    distance (miles)" table where only the one decisive
    (oversubscribed) criterion row ends with two numbers ("<places>
    <distance>") - every other criterion row ends with just the
    places-offered count, so a line-level regex for a trailing
    "int float" pair unambiguously finds the real distance without
    needing pdfplumber's unreliable table extraction on this
    heavily-wrapped layout.
    """
    records = []
    for url in _SOMERSET_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        full_text = ""
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"

        sections = _SOMERSET_SCHOOL_RE.split(full_text)[1:]  # alternating name, body, name, body...
        for name, body in zip(sections[0::2], sections[1::2]):
            matches = _SOMERSET_DISTANCE_LINE_RE.findall(body.split("This information was correct")[0])
            if matches:
                records.append({"school_name": name.strip(), "last_distance_miles": float(matches[-1][1])})
    return records


_DORSET_INDEX_URL = "https://www.dorsetcouncil.gov.uk/w/school-allocations"
_DORSET_LINK_RE = re.compile(r'href="(/documents/d/guest/[^"]+)"')
_DORSET_NAME_RE = re.compile(r"On Time Applications\n(.+?)\n")
_DORSET_DISTANCE_RE = re.compile(r"([\d.]+)\s*Miles\.")


def fetch_dorset() -> list[dict]:
    """Dorset Council publishes one individual "information sheet" PDF
    per oversubscribed school (primary "on-time"/"tr4-ot" and
    secondary "tr7"/"tr9" transfer, both "on-time" and "late" rounds)
    rather than a bulk document - this crawls the index page for the
    current list of documents rather than hardcoding them, filtering
    out "late"-round duplicates of schools already covered by the
    "on-time" round document (find the current index page via
    dorsetcouncil.gov.uk's "school-allocations" page). Clean prose:
    "<School Name>", then "<Criterion> X.XXX Miles." A handful of
    documents (seen for at least one secondary school) are scanned
    images with no extractable text layer - these are silently
    skipped rather than guessed at via OCR.
    """
    print(f"  Downloading index {_DORSET_INDEX_URL}")
    resp = httpx.get(_DORSET_INDEX_URL, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    paths = [p for p in set(_DORSET_LINK_RE.findall(resp.text)) if "late" not in p.lower()]

    records = []
    for path in paths:
        url = f"https://www.dorsetcouncil.gov.uk{path}"
        print(f"  Downloading {url}")
        try:
            pdf_resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
            pdf_resp.raise_for_status()
        except httpx.HTTPError:
            continue
        with pdfplumber.open(io.BytesIO(pdf_resp.content)) as pdf:
            full_text = pdf.pages[0].extract_text() or ""
        name_match = _DORSET_NAME_RE.search(full_text)
        distance_match = _DORSET_DISTANCE_RE.search(full_text)
        if name_match and distance_match:
            records.append({
                "school_name": name_match.group(1).strip(),
                "last_distance_miles": float(distance_match.group(1)),
            })
    return records


_WORCESTERSHIRE_URLS = [
    "https://www.worcestershire.gov.uk/sites/default/files/2026-05/"
    "first_primary_schools_allocation_day_statistics_2026.pdf",
    "https://www.worcestershire.gov.uk/sites/default/files/2026-04/"
    "middle_schools_allocation_day_statistics_2026.pdf",
    "https://www.worcestershire.gov.uk/sites/default/files/2026-02/"
    "high_school_allocation_day_statistics_2026.pdf",
]


def fetch_worcestershire() -> list[dict]:
    """Worcestershire County Council's "Allocation day statistics"
    PDFs (First/Primary + Middle + High), republished each year at new
    URLs under /sites/default/files/YYYY-MM/ (find the current ones
    via worcestershire.gov.uk's "allocation-day-statistics-adjudicator
    -reports-and-resources" page). "District, Name of School, DfE
    Number, PAN, Oversubscribed, Refused, Criterion, Distance (miles)"
    table - column count/order shifts very slightly between the three
    files (a rotated-header artifact, same family of issue as
    Wokingham's), so this uses the same decimal-cell-detection
    approach as Stockport/Peterborough for the distance rather than a
    fixed index; school name is reliably column 1 (District is
    column 0, sometimes blank on continuation rows for the same
    district) in all three.
    """
    records = []
    for url in _WORCESTERSHIRE_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table:
                    continue
                for row in table:
                    if not row or len(row) < 2 or not row[1] or not isinstance(row[1], str):
                        continue
                    distance = None
                    for cell in row:
                        if cell and _STOCKPORT_DECIMAL_RE.match(cell.strip()):
                            distance = float(cell.strip())
                            break
                    if distance is not None:
                        records.append({"school_name": row[1].replace("\n", " ").strip(), "last_distance_miles": distance})
    return records


def fetch_cheshire_east() -> list[dict]:
    """Cheshire East Council's "Breakdown of oversubscription criteria
    allocated" PDF - republished each year at a new URL (find the
    current one via cheshireeast.gov.uk's "previous-allocations.aspx"
    page). Clean table: DfE Number, School, PAN, Total Allocated,
    Lowest Criteria Allocated, Furthest Distance - column header says
    "(miles)" but a subset of rows are actually recorded in metres
    (e.g. "1489", "309", "92.912") rather than miles - confirmed by
    checking what those figures become when divided by 1609.34: each
    one lands on a small, entirely plausible mile value for a
    small/lightly-oversubscribed local primary school, whereas taken
    literally as miles (92, 309, 1489 miles) they're physically
    impossible for a non-boarding UK primary school. Any value over 15
    (a generous upper bound - no legitimate Cheshire primary
    catchment approaches that) is therefore treated as metres and
    converted, rather than either trusted at face value or discarded.
    """
    url = "https://www.cheshireeast.gov.uk/pdf/schools/admissions/2026-criteria-allocated-analysis-primary-lowest-criteria-only.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or len(row) < 6 or not row[1] or not row[-1]:
                    continue
                try:
                    value = float(row[-1].strip())
                except ValueError:
                    continue
                distance = value / _METRES_PER_MILE if value > 15 else value
                records.append({"school_name": row[1].replace("\n", " ").strip(), "last_distance_miles": distance})
    return records


_SUFFOLK_URLS = [
    "https://www.suffolk.gov.uk/asset-library/school-admissions/Directory-of-Schools-in-Suffolk-Primary-2026-2027-1.pdf",
    "https://www.suffolk.gov.uk/asset-library/school-admissions/Directory-of-Schools-in-Suffolk-Secondary-2026-2027-1.pdf",
]
_SUFFOLK_HEADING_RE = re.compile(r"^([A-Za-z][^,\n]{2,80}),.*[A-Z]{1,2}\d[A-Z\d]?\s?\d[A-Z]{2}\s*$")
_SUFFOLK_DISTANCE_RE = re.compile(r"Criterion under which last child admitted:.*?([\d.]+)\s*miles", re.IGNORECASE)


def fetch_suffolk() -> list[dict]:
    """Suffolk County Council's "Directory of Schools" PDFs (Primary +
    Secondary) - a prose directory (school-by-school profile, not a
    table), republished each year at a new URL under /asset-library/
    school-admissions/ (find the current ones via suffolk.gov.uk's
    primary/secondary "apply for a school place" pages). Each school's
    entry is headed by its "<Name>, <Address>, <Postcode>" line
    (detected via a UK postcode regex, since there's no other marker),
    and the decisive line reads "Criterion under which last child
    admitted: <criterion> X.XXX miles" - this scans line-by-line,
    tracking the most recently seen heading so the distance line can
    be attached to the right school even though (unlike every other
    authority in this registry) the layout is continuous prose, not
    row-per-school.
    """
    records = []
    for url in _SUFFOLK_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        full_text = ""
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"

        current = None
        for line in full_text.split("\n"):
            heading_match = _SUFFOLK_HEADING_RE.match(line.strip())
            if heading_match:
                current = heading_match.group(1).strip()
                continue
            distance_match = _SUFFOLK_DISTANCE_RE.search(line)
            if distance_match and current:
                records.append({"school_name": current, "last_distance_miles": float(distance_match.group(1))})
                current = None
    return records


_EAST_SUSSEX_URLS = [
    "https://www.eastsussex.gov.uk/education-learning/schools/apply-for-a-school-place/"
    "apply-for-a-primary-or-junior-school/detailed-school-information-primary-and-junior?print=true",
    "https://www.eastsussex.gov.uk/education-learning/schools/apply-for-a-school-place/"
    "apply-for-a-secondary-school/detailed-school-information-secondary?print=true",
]
_EAST_SUSSEX_TIEBREAK_CAPTION = "Tie-breaker - furthest child allocated a place at the school"
_EAST_SUSSEX_ROW_RE = re.compile(
    r'<td class="govuk-table__cell clean-rich-text" colspan="1">\s*(.+?)\s*</td>\s*'
    r'<td class="govuk-table__cell clean-rich-text govuk-table__cell--numeric[^"]*"\s*colspan="1">'
    r"(.*?)</td>",
    re.DOTALL,
)
_EAST_SUSSEX_DISTANCE_RE = re.compile(r"(\d+)m in category")


def fetch_east_sussex() -> list[dict]:
    """East Sussex County Council publishes its "detailed school
    information" as HTML pages (Primary/Junior + Secondary), each
    split into several alphabetical-range sub-pages in the normal
    view - but the "?print=true" query parameter renders every
    sub-page's content concatenated onto one page, which is far easier
    to scrape than crawling each range separately (find the current
    base URLs via eastsussex.gov.uk's "apply-for-a-primary-or-junior-
    school"/"apply-for-a-secondary-school" sections if this stops
    working). Each alphabetical range has its own "Tie-breaker -
    furthest child allocated a place at the school" table with rows
    like "<School Name> ... XXXXm in category N" - already in metres.
    Rows for schools that weren't oversubscribed just say "All
    preferences allocated" with no metres figure and are naturally
    skipped.
    """
    records = []
    for url in _EAST_SUSSEX_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        for block in resp.text.split(_EAST_SUSSEX_TIEBREAK_CAPTION)[1:]:
            end = block.find("</tbody>")
            section = block[:end] if end > 0 else block[:8000]
            for name, cell in _EAST_SUSSEX_ROW_RE.findall(section):
                match = _EAST_SUSSEX_DISTANCE_RE.search(cell)
                if match:
                    clean_name = name.replace("&nbsp;", "").strip()
                    records.append({
                        "school_name": clean_name,
                        "last_distance_miles": float(match.group(1)) / _METRES_PER_MILE,
                    })
    return records


_WEST_SUSSEX_URLS = [
    "https://www.westsussex.gov.uk/media/e3fn402p/starting_school_stats_2026.pdf",
    "https://www.westsussex.gov.uk/media/p1mlihto/secondary_school_allocation_day_statistics_2026.pdf",
]


def fetch_west_sussex() -> list[dict]:
    """West Sussex County Council's "allocation day statistics" PDFs
    (Starting School/Primary + Secondary) - republished each year at
    new URLs (find the current ones via westsussex.gov.uk's "starting-
    school-places"/"secondary-school-places" pages). Clean table:
    School Name, Places, Offers, Last criteria, Distance (metres) -
    schools that weren't oversubscribed have a long "All applicants...
    were offered a place" sentence instead of a number in the last
    column, which fails the float parse and is correctly skipped.
    """
    records = []
    for url in _WEST_SUSSEX_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table:
                    continue
                for row in table[1:]:
                    if not row or not row[0] or not row[-1]:
                        continue
                    try:
                        distance = float(row[-1].strip())
                    except ValueError:
                        continue
                    name = row[0].replace("\n", " ").strip()
                    records.append({"school_name": name, "last_distance_miles": distance / _METRES_PER_MILE})
    return records


_DURHAM_URLS = [
    "https://durham.gov.uk/article/27947/Primary-school-admissions-individual-school-and-academy-intake-information",
    "https://durham.gov.uk/article/27950/Secondary-school-admissions-individual-school-and-academy-intake-information",
]
_DURHAM_ROW_RE = re.compile(r"<tr>(.*?)</tr>", re.DOTALL)
_DURHAM_CELL_RE = re.compile(r"<td>(.*?)</td>", re.DOTALL)
_DURHAM_DISTANCE_RE = re.compile(r"([\d.]+)\s*[Mm]iles")


def fetch_durham() -> list[dict]:
    """Durham County Council publishes "individual school and academy
    intake information" as plain HTML tables (Primary + Secondary),
    seemingly stable URLs updated in place each year. Each row's last
    cell has prose that differs in exact wording between the two
    documents ("<X.XXX> miles in the distance criterion..." for
    primary, "Last person offered a place lived <X.XXX> miles away
    ..." for secondary, with inconsistent capitalisation of "Miles" in
    a few rows) - rather than handle both phrasings, this just takes
    the first "<number> miles" anywhere in that cell, school name is
    always the second cell. Schools with their own admission
    arrangements ("contact the school directly") have no such number
    and are correctly skipped.
    """
    records = []
    for url in _DURHAM_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        for row in _DURHAM_ROW_RE.findall(resp.text):
            cells = _DURHAM_CELL_RE.findall(row)
            if len(cells) < 2:
                continue
            name = re.sub(r"<[^>]+>", " ", cells[1]).replace("&nbsp;", " ").strip()
            name = re.sub(r"\s+", " ", name)
            if not name:
                continue
            match = _DURHAM_DISTANCE_RE.search(cells[-1])
            if match:
                records.append({"school_name": name, "last_distance_miles": float(match.group(1))})
    return records


_DERBY_URLS = [
    "https://www.derby.gov.uk/media/derbycitycouncil/content/documents/education/schooladmissions/"
    "primary-admissions-handbook-2026-27.pdf",
    "https://www.derby.gov.uk/media/derbycitycouncil/content/documents/education/schooladmissions/"
    "secondary-school-admissions-handbook2026-27.pdf",
]
_DERBY_POSTCODE_RE = re.compile(r"\bDE\d[A-Z\d]?\s?\d[A-Z]{2}\b")
_DERBY_DISTANCE_RE = re.compile(r"Furthest distance offered:\s*([\d.]+)\s*miles")


def fetch_derby() -> list[dict]:
    """Derby City Council's "Admissions Handbook" PDFs (Primary +
    Secondary) - a prose school-by-school prospectus, republished each
    year at a new URL under /media/derbycitycouncil/.../schooladmissions/
    (find the current ones via derby.gov.uk's admissions pages). Each
    school's profile ends with "Admissions Limit: N Furthest distance
    offered: X.XXX miles". The school name heading is detected as the
    line immediately before one containing "Telephone:" and a Derby
    postcode (DEx x-xx) - a few schools have a second subtitle line
    between the name and address (e.g. "Now incorporating <Nursery>")
    which this can't distinguish from the real name, so a handful of
    records get a wrong/partial name and are correctly dropped by the
    fuzzy-match cutoff rather than mismatched.
    """
    records = []
    for url in _DERBY_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        full_text = ""
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"

        lines = full_text.split("\n")
        current = None
        for i, line in enumerate(lines):
            if i + 1 < len(lines) and "Telephone:" in lines[i + 1] and _DERBY_POSTCODE_RE.search(lines[i + 1]):
                current = line.strip()
                continue
            match = _DERBY_DISTANCE_RE.search(line)
            if match and current:
                records.append({"school_name": current, "last_distance_miles": float(match.group(1))})
                current = None
    return records


_LEICESTER_URLS = [
    "https://www.leicester.gov.uk/sites/default/files/2026-02/"
    "breakdown-of-allocations-for-leicester-primary-and-infant-schools-as-of-16-april-2025.pdf",
    "https://schools.leicester.gov.uk/media/8524/"
    "breakdown-of-allocations-for-leicester-city-secondary-schools-as-of-1-march-2023.pdf",
]


def fetch_leicester() -> list[dict]:
    """Leicester City Council's "Breakdown of allocations" PDFs
    (Primary and Infant + Secondary) - republished at a new URL each
    round (find the current ones via leicester.gov.uk's admissions
    pages; the secondary one found is from 2023, the most recent
    working link located - still real published data, just an older
    round than the 2025 primary one). Multiple summary tables appear
    before the real one; only the one whose header contains "Furthest
    Distance" has the data, at whatever column index that table
    happens to place it (found dynamically per table rather than
    hardcoded, since the two files have slightly different column
    sets before it).
    """
    records = []
    for url in _LEICESTER_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table:
                    continue
                header = [h.replace("\n", " ") if h else "" for h in table[0]]
                dist_idx = next((i for i, h in enumerate(header) if "Furthest" in h and "Distance" in h), None)
                if dist_idx is None:
                    continue
                for row in table[1:]:
                    if not row or not row[0] or dist_idx >= len(row) or not row[dist_idx]:
                        continue
                    try:
                        distance = float(row[dist_idx].strip())
                    except ValueError:
                        continue
                    records.append({"school_name": row[0].replace("\n", " ").strip(), "last_distance_miles": distance})
    return records


_SANDWELL_URLS = [
    "https://www.sandwell.gov.uk/downloads/file/388/primary-statistics",
    "https://www.sandwell.gov.uk/downloads/file/389/secondary-statistics",
]


def fetch_sandwell() -> list[dict]:
    """Sandwell Council publishes its admission statistics as a Word
    (.docx) document (not PDF or HTML) - one for primary, one for
    secondary, both at a stable /downloads/file/<id>/ URL (find the
    current ids via sandwell.gov.uk if these ever change). Contains
    several tables; the one we want has the header row "School Name,
    Distance <year1>, Distance <year2>, Distance <year3>" (multiple
    years of history, "N/A" for years the school wasn't oversubscribed
    on distance) - this takes the most recent non-"N/A" year, working
    backwards from the last column, same "most recent real value"
    approach used for Bristol/Solihull's multi-year PDFs.
    """
    records = []
    for url in _SANDWELL_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        doc = docx.Document(io.BytesIO(resp.content))
        for table in doc.tables:
            if not table.rows:
                continue
            header = [c.text.strip() for c in table.rows[0].cells]
            if not header or header[0] != "School Name" or not any("Distance" in h for h in header[1:]):
                continue
            for row in table.rows[1:]:
                cells = [c.text.strip() for c in row.cells]
                if not cells or not cells[0]:
                    continue
                for value in reversed(cells[1:]):
                    try:
                        distance = float(value)
                    except ValueError:
                        continue
                    records.append({"school_name": cells[0], "last_distance_miles": distance})
                    break
    return records


_DUDLEY_NAME_CLEAN_RE = re.compile(r"\s*PAN\s*\d+/\d+\s*[–-]?\s*\d*")


def fetch_dudley() -> list[dict]:
    """Dudley Council's "Primary School Allocations Breakdown" PDF -
    a 3-year time series (2023/2024/2025, most recent first per
    school), republished periodically at a new URL (find the current
    one via dudley.gov.uk's "primary-reception-intake" page). Only the
    first row of each school's 3-row block has the name (a merged
    cell also containing "PAN <year> - <n>", stripped off here); only
    that first row is taken, so this is always the most recent (2025)
    year without needing explicit year comparison. "Furthest Distance
    Admitted" has no unit in the header but is unambiguously in
    metres, not miles (values run into the thousands - a value like
    "8299" is nonsensical as miles for a Dudley primary school but a
    perfectly normal 5.16-mile catchment as metres).
    """
    url = ("https://www.dudley.gov.uk/media/clanjdgu/"
           "dudley-primary-school-allocations-breakdown-from-september-2023-to-september-2025-intakes.pdf")
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table[1:]:
                if not row or not row[0] or len(row) < 12:
                    continue
                name = _DUDLEY_NAME_CLEAN_RE.sub("", row[0].replace("\n", " ")).strip()
                cell = row[11]
                if not cell or cell in ("-", "N/A"):
                    continue
                try:
                    metres = float(cell.replace(",", ""))
                except ValueError:
                    continue
                records.append({"school_name": name, "last_distance_miles": metres / _METRES_PER_MILE})
    return records


def _fetch_afc_borough(url: str) -> list[dict]:
    """Shared parser for the "STARTING PRIMARY SCHOOL... ALLOCATION OF
    PLACES" PDFs used by both Richmond and Kingston upon Thames (both
    administered by the same "Achieving for Children" admissions
    service, hosted on the same rackcdn CDN as RBWM - same underlying
    platform, near-identical layout). Clean table: School, Places,
    EHCP, LAC, Social/medical, Sibling, Child of staff, Distance,
    "Distance of last child offered under criterion 5 in metres" -
    rows reading "All preferences met" or "Overseas" have no number
    and are correctly skipped.
    """
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or not row[0] or not row[-1]:
                    continue
                try:
                    metres = float(row[-1].strip())
                except ValueError:
                    continue
                records.append({"school_name": row[0].replace("\n", " ").strip(), "last_distance_miles": metres / _METRES_PER_MILE})
    return records


def fetch_richmond_upon_thames() -> list[dict]:
    """London Borough of Richmond upon Thames - see
    _fetch_afc_borough() for the shared AfC-platform format. Find the
    current URL via Achieving for Children's site (kr.afcinfo.org.uk)
    or by searching "LBR Notes to parents following Primary <year>
    allocation" - republished each year at a new rackcdn attachment
    ID.
    """
    url = ("https://5f2fe3253cd1dfa0d089-bf8b2cdb6a1dc2999fecbc372702016c.ssl.cf3.rackcdn.com/"
           "uploads/ckeditor/attachments/17717/LBR_Notes_to_parents_following_Primary_2025_allocation.pdf")
    print(f"  Downloading {url}")
    return _fetch_afc_borough(url)


def fetch_kingston_upon_thames() -> list[dict]:
    """Royal Borough of Kingston upon Thames - see
    _fetch_afc_borough() for the shared AfC-platform format (same as
    Richmond). Find the current URL by searching "RBK Notes to parents
    following Primary <year> allocation" - republished each year at a
    new rackcdn attachment ID.
    """
    url = ("https://5f2fe3253cd1dfa0d089-bf8b2cdb6a1dc2999fecbc372702016c.ssl.cf3.rackcdn.com/"
           "uploads/ckeditor/attachments/17727/RBK_Notes_to_parents_following_Primary_2025_allocation.docx.pdf")
    print(f"  Downloading {url}")
    return _fetch_afc_borough(url)


def fetch_southwark() -> list[dict]:
    """London Borough of Southwark's "Allocation of community school
    places by criteria" PDF - republished each year at a new URL
    under /sites/default/files/YYYY-MM/ (find the current one via
    southwark.gov.uk's primary-admissions "admissions-criteria" page -
    note the page itself is served from services.southwark.gov.uk but
    the PDF asset only resolves under www.southwark.gov.uk). Clean
    table, "Furthest distance (metres) offered a school place" at a
    fixed column index; rows reading "ALL APPLICANTS OFFERED A PLACE"
    have no number there and are correctly skipped. Southwark's own
    names are abbreviated to just the distinctive part (e.g. "Crampton"
    for "Crampton Primary School") - too short for the fuzzy matcher to
    find reliably on its own, so " Primary School" is appended before
    matching.
    """
    url = ("https://www.southwark.gov.uk/sites/default/files/2026-03/"
           "Allocation-of-community-school-places-by-criteria-Sept-2025.pdf")
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table[1:]:
                if not row or not row[0] or len(row) < 11 or not row[10]:
                    continue
                try:
                    metres = float(row[10].strip())
                except ValueError:
                    continue
                records.append({
                    "school_name": row[0].strip() + " Primary School",
                    "last_distance_miles": metres / _METRES_PER_MILE,
                })
    return records


_LAMBETH_URLS = [
    "https://www.lambeth.gov.uk/schools-and-education/school-admissions-and-appeals/primary-school-admissions/"
    "how-offers-were-made-lambeth-primary-schools-national-offer-day-16-april-2025/lambeth-voluntary-aided-schools",
    "https://www.lambeth.gov.uk/schools-and-education/school-admissions-and-appeals/primary-school-admissions/"
    "how-offers-were-made-lambeth-primary-schools-national-offer-day-16-april-2025/lambeth-academies",
    "https://www.lambeth.gov.uk/schools-and-education/school-admissions-and-appeals/primary-school-admissions/"
    "how-offers-were-made-lambeth-primary-schools-national-offer-day-16-april-2025/lambeth-foundation-schools",
]
_LAMBETH_HEADING_RE = re.compile(r"<h3[^>]*>(.*?)</h3>")
_LAMBETH_DISTANCE_RE = re.compile(r"Distance for last child[^:]{0,60}:\s*(\d+\.?\d*)")


def fetch_lambeth() -> list[dict]:
    """London Borough of Lambeth publishes its "how offers were made"
    results as plain HTML prose pages (not PDFs), split by admission-
    authority type (voluntary aided / academies / foundation schools -
    community schools apparently not published this way, or not
    oversubscribed), each republished at the same URL pattern each
    year with just the date in the path changing (find current ones
    via lambeth.gov.uk's primary-admissions pages). Uses "&nbsp;"
    entities instead of literal spaces throughout, which breaks any
    regex expecting whitespace unless replaced first. Some banded
    schools (Foundation/Open places) report two separate "Distance for
    last child" figures - this takes the larger, consistent with the
    same choice made for other multi-criterion authorities (Brent,
    Surrey). One occurrence (Iqra Primary School) is mislabelled
    "miles" in the source when the value is unambiguously in metres
    (consistent with every other entry's magnitude, and no school
    admits from 1,303 miles away) - all values are therefore always
    treated as metres regardless of the literal unit word.
    """
    records = []
    for url in _LAMBETH_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        text = resp.text.replace("&nbsp;", " ")
        current = None
        for chunk in re.split(r"(<h3[^>]*>.*?</h3>)", text):
            heading_match = _LAMBETH_HEADING_RE.match(chunk)
            if heading_match:
                current = re.sub("<[^>]+>", "", heading_match.group(1)).strip()
                continue
            if current:
                distances = [float(d) for d in _LAMBETH_DISTANCE_RE.findall(chunk)]
                if distances:
                    records.append({"school_name": current, "last_distance_miles": max(distances) / _METRES_PER_MILE})
                current = None
    return records


def fetch_waltham_forest() -> list[dict]:
    """London Borough of Waltham Forest's "Starting Primary School"
    brochure PDF - a large prospectus, republished each year at a new
    URL (find the current one via walthamforest.gov.uk's primary
    admissions page). Most of the document uses rotated column
    headers that extract as reversed/garbled text, but the specific
    "Cut off distances for the past 3 years" table (found by header
    text "School, <year1>, <year2>, <year3>") is a clean simple table,
    already in miles - takes the most recent non-blank year, working
    backwards from the last column, same "most recent real value"
    approach used for other multi-year sources (Bristol, Solihull,
    Sandwell).
    """
    url = "https://www.walthamforest.gov.uk/sites/default/files/2025-08/1012501%20-%20WF_PRIMARY%20BROCHURE%202026%20-%20WEB.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table or not table[0] or table[0][0] != "School":
                continue
            for row in table[1:]:
                if not row or not row[0]:
                    continue
                for value in reversed(row[1:]):
                    if value and value.strip():
                        try:
                            records.append({"school_name": row[0].strip(), "last_distance_miles": float(value.strip())})
                        except ValueError:
                            pass
                        break
    return records


def fetch_lewisham() -> list[dict]:
    """London Borough of Lewisham's "Applying to start primary
    school" brochure PDF - republished each year at the same stable
    URL (find via lewisham.gov.uk's primary admissions page if it
    ever moves). Most of the document's tables use rotated column
    headers that extract as vertically-garbled text, but the "Name of
    school" table (only 2 pages, community schools) is otherwise
    clean - the decisive metres figure is found via the same
    decimal-cell-detection approach as Stockport/Peterborough rather
    than a fixed column index. Lewisham's own names are abbreviated
    (e.g. "Adamsrill" for "Adamsrill Primary School"), so " Primary
    School" is appended before matching, same fix as Southwark.
    """
    url = ("https://lewisham.gov.uk/-/media/services/education/schools/primary-school/"
           "applying-to-start-primary-school-in-september.pdf")
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table or not table[0] or table[0][0] != "Name of school":
                continue
            for row in table[1:]:
                if not row or not row[0]:
                    continue
                distance = None
                for cell in row[1:]:
                    if cell and _STOCKPORT_DECIMAL_RE.match(cell.strip()):
                        distance = float(cell.strip())
                        break
                if distance is not None:
                    records.append({
                        "school_name": row[0].strip() + " Primary School",
                        "last_distance_miles": distance / _METRES_PER_MILE,
                    })
    return records


def fetch_merton() -> list[dict]:
    """London Borough of Merton's "Admissions and appeals data for
    recent years" PDF - a single large document going back several
    years, republished at a stable URL (find current one via
    merton.gov.uk's "recent-years" admissions page if it moves). Most
    recent secondary round is page index 2 (dated 3/3/25, right after
    the "Transfer to secondary school" section heading); most recent
    primary round is pages 9-11 (dated 16/4/2024 - the primary round
    is a full year further behind than secondary in this document,
    genuinely the latest published at time of writing). Each row can
    show two distance figures (initial + "second round" after
    appeals/late applications, sometimes "TBC" if not yet published);
    this takes the max of whichever are real numbers, same policy as
    other multi-round authorities. Merton's own primary names are
    abbreviated (e.g. "Hillcross" for "Hillcross Primary School"), so
    " Primary School" is appended only for the primary pages (the
    secondary page's names are already complete).
    """
    url = "https://www.merton.gov.uk/system/files/admissions_and_appeals_data_for_recent_years_-_table.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page_index, suffix in [(2, ""), (9, " Primary School"), (10, " Primary School"), (11, " Primary School")]:
            table = pdf.pages[page_index].extract_table()
            if not table:
                continue
            for row in table:
                if not row or not row[0]:
                    continue
                name = row[0].replace("\n", " ").strip()
                distances = []
                for cell in row[1:]:
                    if cell and "." in cell:
                        try:
                            distances.append(float(cell.strip()))
                        except ValueError:
                            pass
                if distances:
                    records.append({"school_name": name + suffix, "last_distance_miles": max(distances) / _METRES_PER_MILE})
    return records


_GREENWICH_DISTANCE_RE = re.compile(r"([\d.]+)")


def fetch_greenwich() -> list[dict]:
    """Royal Borough of Greenwich's "Primary School admissions data
    and statistics" Excel workbook - republished each year at a new
    URL under /sites/default/files/YYYY-MM/ (find the current one via
    royalgreenwich.gov.uk's downloads page, ID 1191 as of writing - no
    secondary equivalent was found). Uniquely among every source in
    this registry, it publishes the school's own URN directly (column
    "URN & DFE Numbers", formatted "<URN> & <DfE No>") rather than
    relying on name matching - so records here carry a "urn" key that
    build_records() uses directly (still checked against this
    authority's real URNs rather than trusted blindly). "Last distance
    offered in metres" is sometimes annotated (e.g. "21437.68 (Open
    Band)" for faith schools with banded admission) - only the leading
    number is taken. Section-header rows ("Planning Area 1") have no
    PAN and are naturally skipped.
    """
    url = "https://www.royalgreenwich.gov.uk/sites/default/files/2025-04/Primary_Stats_for_Website_2024.xlsx"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb[wb.sheetnames[0]]

    records = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[1] or not row[2] or not row[20]:
            continue
        urn_match = re.match(r"\s*(\d+)", str(row[1]))
        distance_match = _GREENWICH_DISTANCE_RE.search(str(row[20]))
        if urn_match and distance_match:
            records.append({
                "school_name": (row[0] or "").strip(),
                "urn": int(urn_match.group(1)),
                "last_distance_miles": float(distance_match.group(1)) / _METRES_PER_MILE,
            })
    return records


_SUTTON_PRIMARY_URL = (
    "https://www.sutton.gov.uk/schools-and-learning/school-admissions/"
    "national-offer-day-primary-schools/primary-school-allocation-information"
)
_SUTTON_SECONDARY_URL = "https://www.sutton.gov.uk/sites/default/files/2026-04/TSS%20Guidance%20Booklet%202026v2.pdf"
_SUTTON_ROW_RE = re.compile(r"<tr><td>(.*?)</td><td>[^<]*</td><td>[^<]*</td><td>([^<]*)</td></tr>")
_SUTTON_NAME_SUFFIX_RE = re.compile(r"\s*\((Academy|Foundation|Voluntary-aided|Voluntary-controlled)\)\s*$")
_SUTTON_DECIMAL_RE = re.compile(r"[\d,]+\.\d+")


def fetch_sutton() -> list[dict]:
    """London Borough of Sutton publishes its primary allocation
    figures as a live HTML table (not a PDF) on its "primary-school-
    allocation-information" page, and its secondary figures inside the
    "Transfer to Secondary School" guidance booklet PDF, republished
    each year at a new URL under /sites/default/files/YYYY-MM/ (find
    the current one via sutton.gov.uk's secondary transfer page).
    Names carry a trailing admission-authority-type annotation (e.g.
    "Avenue Primary (Academy)") stripped before matching. Distances
    use comma thousands-separators ("2,697.92") which float() can't
    parse directly; a school with two sites (e.g. "Cheam High School
    (Worcester Park Places)") reports two figures on separate lines
    within one cell - takes the larger, consistent with other
    multi-figure sources.
    """
    records = []

    print(f"  Downloading {_SUTTON_PRIMARY_URL}")
    resp = httpx.get(_SUTTON_PRIMARY_URL, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    for name, cell in _SUTTON_ROW_RE.findall(resp.text):
        name = _SUTTON_NAME_SUFFIX_RE.sub("", name).strip()
        matches = _SUTTON_DECIMAL_RE.findall(cell.replace(",", ""))
        if matches:
            records.append({"school_name": name, "last_distance_miles": max(float(m) for m in matches) / _METRES_PER_MILE})

    print(f"  Downloading {_SUTTON_SECONDARY_URL}")
    resp = httpx.get(_SUTTON_SECONDARY_URL, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table or not table[0] or "Distance" not in str(table[0][-1]):
                continue
            for row in table[1:]:
                if not row or not row[0]:
                    continue
                matches = _SUTTON_DECIMAL_RE.findall((row[-1] or "").replace(",", ""))
                if matches:
                    name = row[0].replace("\n", " ").strip()
                    records.append({"school_name": name, "last_distance_miles": max(float(m) for m in matches) / _METRES_PER_MILE})
    return records


_HARINGEY_URLS = [
    "https://haringey.gov.uk/schools-learning/schools/school-admissions/how-school-place-offers-were-made/"
    "cutoff-distance-school-last-child-offered-place/"
    "primary-schools-distance-school-last-child-offered-place-national-offer-day",
    "https://haringey.gov.uk/schools-learning/schools/school-admissions/how-school-place-offers-were-made/"
    "cutoff-distance-school-last-child-offered-place/"
    "distance-school-last-child-offered-place-1-september",
]
_HARINGEY_ROW_RE = re.compile(r"<tr>\s*<td>(.*?)</td>((?:\s*<td>.*?</td>)+)\s*</tr>", re.DOTALL)
_HARINGEY_CELL_RE = re.compile(r"<td>(.*?)</td>")


def fetch_haringey() -> list[dict]:
    """Haringey Council publishes live HTML tables (not PDFs) of
    "Distance (in miles) from school of last child offered a place" -
    one for primary (national offer day, columns newest year first)
    and one for secondary (1 September, i.e. after the summer's late
    applications/appeals - same newest-first column order), both
    already in miles across several years. Takes the first (most
    recent) non-"N/A" column per row rather than always the newest
    year, since some schools' most recent year has no distance figure
    (not oversubscribed that year).
    """
    records = []
    for url in _HARINGEY_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        text = resp.text.replace("&nbsp;", " ")
        start = text.find("<tbody>")
        end = text.find("</tbody>")
        if start < 0 or end < 0:
            continue
        for name, cells in _HARINGEY_ROW_RE.findall(text[start:end]):
            clean_name = re.sub("<[^>]+>", "", name).strip()
            for value in _HARINGEY_CELL_RE.findall(cells):
                value = value.strip()
                if value and value.upper() not in ("N/A", "ALL"):
                    try:
                        records.append({"school_name": clean_name, "last_distance_miles": float(value)})
                    except ValueError:
                        pass
                    break
    return records


_CALDERDALE_URLS = [
    "https://dataworks.calderdale.gov.uk/download/vq9q6/t8y/Primary%20Preferences-Allocations%20by%20School%202026.csv",
    "https://dataworks.calderdale.gov.uk/download/v8357/127/Secondary%20Preferences-Allocations%20by%20School%202026.csv",
]
_CALDERDALE_DISTANCE_RE = re.compile(r"([\d.]+)")


def fetch_calderdale() -> list[dict]:
    """Calderdale Council publishes real open-data CSVs (Primary +
    Secondary preferences/allocations by school) on its "Calderdale
    Data Works" open data portal, one file per year since 2015 -
    republished each year at a new download ID (find the current ones
    via the dataset pages: dataworks.calderdale.gov.uk/dataset/vq9q6
    and .../v8357). "Distance of furthest pupil allocated a place (in
    miles)" is already in miles; a few rows carry an annotation like
    "0.351 (within catchment)" - only the leading number is taken.
    Blank cells (school not oversubscribed on distance) are correctly
    skipped.
    """
    records = []
    for url in _CALDERDALE_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        text = resp.content.decode("utf-8-sig", errors="replace")
        for line in text.splitlines()[1:]:
            if not line.strip():
                continue
            cells = next(csv.reader([line]))
            if len(cells) < 10 or not cells[0] or not cells[9].strip():
                continue
            match = _CALDERDALE_DISTANCE_RE.search(cells[9])
            if match:
                records.append({"school_name": cells[0].strip(), "last_distance_miles": float(match.group(1))})
    return records


_WALSALL_URLS = [
    "https://go.walsall.gov.uk/sites/default/files/2026-04/Reception%20Offers%202026_0.pdf",
    "https://go.walsall.gov.uk/sites/default/files/2026-04/Year%203%20Offers%202026_Nicholas%20Gill.pdf",
]


def fetch_walsall() -> list[dict]:
    """Walsall Council's "Reception Offers" and "Year 3 Offers" PDFs -
    republished each year at a new URL under /sites/default/files/
    YYYY-MM/ (find the current ones via go.walsall.gov.uk's primary
    admissions page). Rotated "Furthest Distance" column header
    extracts as reversed text, and the two files have a different
    column count (Year 3 has no Pupil Premium column), so the decisive
    figure is found via the same decimal-cell-detection approach as
    Stockport/Peterborough rather than a fixed index. Already in
    miles; "N/A"/"n/a" rows (not oversubscribed on distance) are
    naturally skipped.
    """
    records = []
    for url in _WALSALL_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table:
                    continue
                for row in table:
                    if not row or not row[0] or not isinstance(row[0], str) or not row[1] or not row[1].isdigit():
                        continue
                    distance = None
                    for cell in row[2:]:
                        if cell and _STOCKPORT_DECIMAL_RE.match(cell.strip()):
                            distance = float(cell.strip())
                            break
                    if distance is not None:
                        records.append({"school_name": row[0].replace("\n", " ").strip(), "last_distance_miles": distance})
    return records


def fetch_oldham() -> list[dict]:
    """Oldham Council's "Primary admissions summary of last place
    offered" PDF - republished each year at a new URL (find the
    current one via oldham.gov.uk, no secondary equivalent found).
    Clean table: School (prefixed "(Oldham) ", stripped here), PAN,
    Criteria, Distance - already in miles. "*" marks schools not
    oversubscribed on distance and is correctly skipped.
    """
    url = "https://www.oldham.gov.uk/download/downloads/id/8161/2025_primary_admissions_summary_of_last_place_offered.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table[1:]:
                if not row or not row[0] or not row[-1]:
                    continue
                try:
                    distance = float(row[-1].strip())
                except ValueError:
                    continue
                name = row[0].replace("\n", " ").replace("(Oldham)", "").strip()
                records.append({"school_name": name, "last_distance_miles": distance})
    return records


_WIGAN_INDEX_URLS = [
    "https://www.wigan.gov.uk/resident/education/schools/school-admissions/primary-schools.aspx",
    "https://www.wigan.gov.uk/resident/education/schools/School-Admissions/Secondary-Schools.aspx",
]
_WIGAN_LINK_RE = re.compile(r'href="(/Docs/PDF/Resident/Education/Schools/Admissions/[^"]+\.pdf)"')
_WIGAN_DISTANCE_RE = re.compile(r"living\s*([\d.]+)\s*miles from the school")


def fetch_wigan() -> list[dict]:
    """Wigan Council publishes one individual "how places were
    allocated" PDF per school (both primary and secondary), listed on
    two index pages that this crawls for the current list rather than
    hardcoding them (find the current index pages via wigan.gov.uk's
    "school-admissions" section if the URL slugs change; excludes the
    generic booklets and appeals-info PDF that also appear in the same
    link list). Prose format: school name is the document's first
    line, and the decisive line reads "...living X.XXX miles from the
    school." Schools not oversubscribed have no such sentence and are
    naturally skipped.
    """
    records = []
    for index_url in _WIGAN_INDEX_URLS:
        print(f"  Downloading index {index_url}")
        resp = httpx.get(index_url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        paths = [p for p in set(_WIGAN_LINK_RE.findall(resp.text))
                 if "Booklet" not in p and "Appeals" not in p]
        for path in paths:
            url = f"https://www.wigan.gov.uk{path}"
            print(f"  Downloading {url}")
            try:
                pdf_resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
                pdf_resp.raise_for_status()
            except httpx.HTTPError:
                continue
            with pdfplumber.open(io.BytesIO(pdf_resp.content)) as pdf:
                full_text = pdf.pages[0].extract_text() or ""
            lines = full_text.split("\n")
            if not lines:
                continue
            name = lines[0].strip()
            match = _WIGAN_DISTANCE_RE.search(full_text)
            if match:
                records.append({"school_name": name, "last_distance_miles": float(match.group(1))})
    return records


def fetch_tameside() -> list[dict]:
    """Tameside's Primary and Secondary "allocation statistics" PDFs
    both publish the school's own real URN directly (Primary: a "URN"
    column; Secondary: a "School unique ref number" column) - matched
    directly like Greenwich, bypassing fuzzy name matching entirely
    (name cells are frequently garbled by pdfplumber merging wrapped
    multi-line cells, but the URN + Distance columns stay aligned
    regardless). One secondary row has its URN corrupted to "E 106270"
    by a line-wrap artifact - handled by extracting the digit run
    rather than trusting the whole cell.
    """
    records = []

    primary_url = "https://www.tameside.gov.uk/documents/d/guest/primary-full-stats-2024_1-pdf"
    resp = httpx.get(primary_url, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if len(row) < 18 or not row[2] or not row[2].isdigit() or not row[17]:
                    continue
                try:
                    distance = float(row[17])
                except ValueError:
                    continue
                records.append({"urn": int(row[2]), "school_name": row[3] or "", "last_distance_miles": distance})

    secondary_url = "https://www.tameside.gov.uk/documents/d/guest/secondary-2025-pdf"
    resp = httpx.get(secondary_url, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if len(row) < 21 or not row[1] or not row[20]:
                    continue
                urn_match = re.search(r"\d{5,7}", row[1])
                if not urn_match:
                    continue
                try:
                    distance = float(row[20])
                except ValueError:
                    continue
                records.append({"urn": int(urn_match.group()), "school_name": row[0] or "", "last_distance_miles": distance})

    return records


def fetch_gloucestershire() -> list[dict]:
    """Gloucestershire's Primary, Junior and Secondary "allocation day
    statistics" PDFs - clean tables with "Furthest distance allocated
    (miles)" as a fixed column. Rows that weren't oversubscribed show
    "N/A" (skipped); distances beyond 20 miles are deliberately
    redacted by the council to ">20" for data-protection reasons
    (skipped rather than guessed at). The "School DfE no." column is
    the LOCAL 4-digit establishment number, not the national URN, so
    matching is by (fuzzy) name like most other authorities.
    """
    urls = [
        "https://www.gloucestershire.gov.uk/media/ctgfusdi/primary-allocation-day-statistics-2025-1.pdf",
        "https://gloucestershire.gov.uk/media/0rzhtukd/junior-allocation-day-statistics-2025.pdf",
        "https://www.gloucestershire.gov.uk/media/v1jdwr0r/secondary-allocation-day-statistics-2025.pdf",
    ]
    records = []
    for url in urls:
        resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table:
                    continue
                for row in table:
                    if len(row) < 6 or not row[1] or not row[5]:
                        continue
                    try:
                        distance = float(row[5])
                    except ValueError:
                        continue
                    records.append({"school_name": row[1], "last_distance_miles": distance})
    return records


_WARWICKSHIRE_URLS = [
    "https://api.warwickshire.gov.uk/documents/WCCC-1990003847-4065",  # Reception 2025
    "https://api.warwickshire.gov.uk/documents/WCCC-1990003847-4066",  # Junior 2025
    "https://api.warwickshire.gov.uk/documents/WCCC-1990003847-3909",  # Secondary 2025
]


def fetch_warwickshire() -> list[dict]:
    """Warwickshire's Reception/Junior/Secondary "breakdown of offers"
    spreadsheets. Unlike most sources, EVERY row has a "Distance"
    figure regardless of whether that place was actually decided by
    proximity - undersubscribed schools that admitted a distant
    out-of-area applicant report that applicant's (irrelevant, huge)
    distance too, and the "Last Offer Made Criterion" text doesn't
    reliably distinguish this case (school-specific naming, and even
    some "In Priority Area" rows have 90+ mile outliers). Rather than
    trust or parse that free-text criterion column, distances over 20
    miles are treated as not representative of a real catchment and
    dropped (~7% of rows) - the same plausibility-cap approach used
    for other noisy sources (see Peterborough, Wokingham).
    """
    records = []
    for url in _WARWICKSHIRE_URLS:
        resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0] or row[4] is None:
                continue
            try:
                distance = float(row[4])
            except (TypeError, ValueError):
                continue
            if distance > 20:
                continue
            records.append({"school_name": row[0], "last_distance_miles": distance})
    return records


_STAFFORDSHIRE_URLS = [
    "https://www.staffordshire.gov.uk/schools-and-learning/school-admissions/admission-oversubscribed-schools/summary-september-2025-0",
    "https://www.staffordshire.gov.uk/schools-and-learning/school-admissions/admission-oversubscribed-schools/summary-september-2026",
]
_STAFFS_TABLE_RE = re.compile(r"<table.*?>(.*?)</table>", re.DOTALL)
_STAFFS_HEADER_RE = re.compile(r"<th[^>]*>(.*?)</th>", re.DOTALL)
_STAFFS_ROW_RE = re.compile(r"<tr>(.*?)</tr>", re.DOTALL)
_STAFFS_CELL_RE = re.compile(r"<td>(.*?)</td>", re.DOTALL)


def _staffs_clean(text: str) -> str:
    text = re.sub(r"<[^>]+>", " ", text).replace("&nbsp;", " ").replace("&amp;", "&")
    return re.sub(r"\s+", " ", text).strip()


def fetch_staffordshire() -> list[dict]:
    """Staffordshire's "summary of admission to oversubscribed
    schools" pages (Primary 2025 + Secondary 2026 - the most recent
    of each found), one HTML table per school with a "Furthest
    distance (miles)" column. The exact set of preceding criteria
    columns (Sibling/Catchment/Staff child/Pupil Premium/etc.) varies
    per school, so the distance column's index is located dynamically
    from each table's own header rather than assumed fixed. Schools
    not oversubscribed on distance show a blank or "N/A" cell there
    and are skipped.
    """
    records = []
    for url in _STAFFORDSHIRE_URLS:
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        for table_html in _STAFFS_TABLE_RE.findall(resp.text):
            headers = [_staffs_clean(h) for h in _STAFFS_HEADER_RE.findall(table_html)]
            dist_idx = next((i for i, h in enumerate(headers) if "furthest distance" in h.lower()), None)
            if dist_idx is None:
                continue
            for row in _STAFFS_ROW_RE.findall(table_html):
                cells = _STAFFS_CELL_RE.findall(row)
                if len(cells) != len(headers):
                    continue
                name = _staffs_clean(cells[0])
                dist_text = _staffs_clean(cells[dist_idx])
                if not name or not dist_text:
                    continue
                try:
                    distance = float(dist_text)
                except ValueError:
                    continue
                records.append({"school_name": name, "last_distance_miles": distance})
    return records


def fetch_kirklees() -> list[dict]:
    """Kirklees' Reception "by preference and criteria" PDF - several
    tables per document (Community/Voluntary Controlled schools, Own
    Admission Authority schools, and a preference-summary table with
    no distance column at all), with a differing number of criteria
    columns between the first two. Rather than assume a fixed layout,
    only tables whose header row mentions "distance" are used, and
    within those the PAN column (index 1) must be a plain digit to
    skip header/section-title rows. Distances are in metres. No
    secondary/Year 7 equivalent was found at a stable URL.
    """
    url = "https://www.kirklees.gov.uk/beta/admissions/pdf/reception-by-preference-and-criteria-25.pdf"
    resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                if not table or not any(h and "distance" in h.lower() for h in table[0]):
                    continue
                for row in table[1:]:
                    if len(row) < 2 or not row[1] or not row[1].isdigit():
                        continue
                    name, dist_text = row[0], row[-1]
                    if not name or not dist_text:
                        continue
                    try:
                        distance_m = float(dist_text)
                    except ValueError:
                        continue
                    records.append({"school_name": name, "last_distance_miles": distance_m / _METRES_PER_MILE})
    return records


_CHESHIRE_WEST_URLS = [
    "https://www.cheshirewestandchester.gov.uk/asset-library/cwc-primary-guide-2025-26-online-final-version-mar-20251.pdf",
    "https://www.cheshirewestandchester.gov.uk/asset-library/secondary-school-guide-2025-26-final-online12.pdf",
]


_CHESHIRE_WEST_POSTCODE_RE = re.compile(r"[A-Z]{1,2}\d[A-Z\d]?\s+\d[A-Z]{2}$")


def _cheshire_west_name(cell: str) -> str:
    lines = [line.strip() for line in cell.split("\n") if line.strip()]
    if not lines:
        return ""
    postcode_idx = next((i for i, line in enumerate(lines) if _CHESHIRE_WEST_POSTCODE_RE.search(line)), None)
    if not postcode_idx:
        return lines[0]
    # The address is 1 or 2 lines ending at the postcode line - a
    # 2-line address (street, then town + postcode) is told apart from
    # a 2-line school name by whether the preceding line has a comma
    # (an address fragment) rather than being part of the name, but
    # never mistake line 0 itself for an address line even if the
    # school's own name happens to contain a comma (e.g. "The County
    # High School, Leftwich").
    addr_start = postcode_idx
    if postcode_idx - 1 > 0 and "," in lines[postcode_idx - 1]:
        addr_start = postcode_idx - 1
    return " ".join(lines[:addr_start]) or lines[0]


def fetch_cheshire_west_and_chester() -> list[dict]:
    """Cheshire West and Chester's Primary and Secondary admissions
    guide PDFs have a per-school table row with a "Lowest criteria"
    and "Furthest distance" column - but pdfplumber extracts this
    section of every page as right-to-left, so BOTH the criteria label
    ("Distance" -> "ecnatsiD") and the distance figure itself
    (e.g. "1.049" -> "940.1") come out character-reversed, not just the
    column headers (the more common case elsewhere in this script).
    Only rows whose reversed criteria label is exactly "Distance" are
    used; other rows (a named criterion number, "N/A", or blank) mean
    the school wasn't oversubscribed on distance and are skipped. The
    school name can wrap across 1-2 lines before the address line
    (identified by the following "Tel:" line), so both are joined.
    """
    records = []
    for url in _CHESHIRE_WEST_URLS:
        resp = httpx.get(url, timeout=90, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table:
                    continue
                for row in table:
                    if len(row) < 3:
                        continue
                    criterion = (row[-3] or "").replace("\n", "").strip()
                    if criterion.lower() != "ecnatsid":
                        continue
                    value = (row[-2] or "").strip()
                    if not value:
                        continue
                    try:
                        distance = float(value[::-1])
                    except ValueError:
                        continue
                    name = _cheshire_west_name(row[1] or "")
                    if name:
                        records.append({"school_name": name, "last_distance_miles": distance})
    return records


def fetch_bristol() -> list[dict]:
    """Bristol's "furthest distance table" - one row per primary
    school with a column per year back to 2020 (most recent last,
    2026), distance in kilometres. Non-numeric markers ("D" = not
    needed as a tie-break, "O" = faith/random allocation rather than
    distance, "N" = not open) mean that year isn't usable, so this
    scans backwards from the most recent year and takes the first
    column that parses as a plain number. A handful of cells in older
    year-columns are corrupted by an overlapping-text PDF rendering
    glitch (e.g. "0.68 6", "D7") - these simply fail to parse and are
    skipped in the backwards scan like any other non-numeric year.
    """
    url = "https://www.bristol.gov.uk/files/documents/3382-furthest-distance-table/file"
    resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or not row[0] or len(row) < 2:
                    continue
                name = row[0].strip()
                if not name or name.lower().startswith("name of school"):
                    continue
                for cell in reversed(row[1:]):
                    if not cell:
                        continue
                    try:
                        distance_km = float(cell.strip())
                    except ValueError:
                        continue
                    records.append({"school_name": name, "last_distance_miles": distance_km / 1.60934})
                    break
    return records


_NORTH_YORKSHIRE_URLS = [
    "https://hub.datanorthyorkshire.org/dataset/8e128b63-0967-43bb-929d-a4058e30f1e3/resource/94bf88ea-6c2f-4b69-80a5-09e990b2429b/download/2025-26-primary-statistics-school-admissions.xlsx",
    "https://hub.datanorthyorkshire.org/dataset/8e128b63-0967-43bb-929d-a4058e30f1e3/resource/33e95fc2-8373-4fe0-80cd-dfccbddcfaa7/download/2025-26-secondary-statistics-school-admissions.xlsx",
]
_NORTH_YORKSHIRE_DISTANCE_RE = re.compile(r"([\d.]+)")


def fetch_north_yorkshire() -> list[dict]:
    """North Yorkshire's open-data Primary/Secondary admissions
    spreadsheets (Data North Yorkshire hub) - the distance column is
    free text like "Out of Catchment Distance - 3.433 miles" rather
    than a plain number, so the first decimal number in the cell is
    extracted. "N/A" (not oversubscribed) and "Contact the school"
    (own admission authority) rows have no number and are skipped.
    """
    records = []
    for url in _NORTH_YORKSHIRE_URLS:
        resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[1] or not row[8]:
                continue
            match = _NORTH_YORKSHIRE_DISTANCE_RE.search(str(row[8]))
            if not match:
                continue
            records.append({"school_name": row[1], "last_distance_miles": float(match.group(1))})
    return records


_READING_PRIMARY_URL = "https://brighterfuturesforchildren.org/wp-content/uploads/2025/04/Primary-Junior-Allocations-for-All-Schools-2025.pdf"
_READING_SECONDARY_URL = "https://brighterfuturesforchildren.org/wp-content/uploads/2025/03/Secondary-Places-2025-Table.pdf"
_READING_BRACKET_DISTANCE_RE = re.compile(r"\((\d+\.\d+)\)")
_READING_SECONDARY_NAME_RE = re.compile(r"([A-Z][^\n]{2,80}?)\n(?:Category )?Admission Number")
_READING_SECONDARY_DISTANCE_RE = re.compile(r"([\d.]+) miles from")


def fetch_reading() -> list[dict]:
    """Reading's (published by its outsourced children's-services
    trust, Brighter Futures for Children) Primary/Junior allocations
    table embeds the tie-break distance in brackets within whichever
    oversubscription-category cell it happened to occur in (e.g.
    "17 (0.331)", vs a non-distance bracket like "5 (1)" for an EYPP
    count) - searching the whole row for a bracketed DECIMAL number is
    the only reliable way to find it, since decimals never appear for
    plain headcounts. The Secondary document is a different, prose
    layout with no distinct table per school (short entries are
    packed onto shared pages) - schools are split apart by matching
    the school-name line that immediately precedes "Admission Number",
    then a "<X.XXX> miles from" sentence is searched for within each
    school's block of text.
    """
    records = []

    resp = httpx.get(_READING_PRIMARY_URL, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or not row[0]:
                    continue
                name = row[0].replace("\n", " ").strip()
                if not name or "oversubscription" in name.lower() or "admissions to" in name.lower() or name.lower() == "name of school":
                    continue
                joined = " ".join(cell for cell in row if cell)
                matches = _READING_BRACKET_DISTANCE_RE.findall(joined)
                if matches:
                    records.append({"school_name": name, "last_distance_miles": float(matches[-1])})

    resp = httpx.get(_READING_SECONDARY_URL, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    full_text = ""
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"
    name_matches = list(_READING_SECONDARY_NAME_RE.finditer(full_text))
    for i, m in enumerate(name_matches):
        name = m.group(1).rstrip(".").strip()
        start = m.end()
        end = name_matches[i + 1].start() if i + 1 < len(name_matches) else len(full_text)
        distance_match = _READING_SECONDARY_DISTANCE_RE.search(full_text[start:end])
        if distance_match:
            records.append({"school_name": name, "last_distance_miles": float(distance_match.group(1))})

    return records


_BURY_URLS = [
    "https://www.bury.gov.uk/asset-library/historical-primary-20251.pdf",
    "https://www.bury.gov.uk/asset-library/historical-community-secondary-schools-allocation-information2.pdf",
]
_BURY_NAME_RE = re.compile(r"\n([A-Za-z][^\n]{2,60})\nPublished")


def fetch_bury() -> list[dict]:
    """Bury's "historical information" PDFs (Primary, and Community
    Secondary - the separate Voluntary Aided secondary one has no
    distance column at all, faith/preference-based only, so it's
    skipped) list 5-6 years per school, each year on its own line with
    a distance figure somewhere in it (its column position differs
    between the two documents), or "All"/"-" for a year not
    oversubscribed on distance. Schools are split apart by matching
    the name line immediately before a "Published" header line; within
    each school's block, the most recent year (scanning backwards)
    whose line contains an actual decimal number is used - this also
    naturally skips a row mangled by an unrelated PDF text-wrapping
    glitch (seen on one school, "The Derby") that dropped its real
    distance onto a separate line, since the visible integer-only
    remainder has no decimal for this to match. Primary school names
    in that document are bare ("Cams Lane"), so " Primary School" is
    appended before matching - secondary names are already given in
    full ("Hazel Wood" -> "Hazel Wood High School" fuzzy-matches fine
    without one).
    """
    records = []
    for url in _BURY_URLS:
        is_primary = url is _BURY_URLS[0]
        resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        full_text = ""
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"
        matches = list(_BURY_NAME_RE.finditer(full_text))
        for i, m in enumerate(matches):
            name = m.group(1).strip()
            start = m.end()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(full_text)
            block = full_text[start:end]
            for year in range(2026, 2020, -1):
                year_match = re.search(rf"^{year}\s+(.*)$", block, re.MULTILINE)
                if not year_match:
                    continue
                # Distance is the only decimal-formatted value on the line (PAN,
                # preference/category counts, appeals and totals are all plain
                # integers) - its column position differs between the Primary and
                # Secondary documents, so this searches for the decimal rather than
                # assuming an index. This also naturally excludes a row mangled by
                # a PDF text-wrapping glitch (seen on one school, "The Derby") that
                # dropped its real distance onto a separate line, since the
                # visible integer-only remainder has no "." for this to match.
                distance_match = re.search(r"(\d+\.\d+)", year_match.group(1))
                if not distance_match:
                    continue
                full_name = f"{name} Primary School" if is_primary else name
                records.append({"school_name": full_name, "last_distance_miles": float(distance_match.group(1))})
                break
    return records


_BOLTON_CATEGORY_URLS = [
    "https://www.bolton.gov.uk/directory/4/school-directory/category/12",  # Primary
    "https://www.bolton.gov.uk/directory/4/school-directory/category/15",  # Secondary
]
_BOLTON_ROW_RE = re.compile(
    r"<td>last distance offered</td>((?:\s*<td[^>]*>[^<]*</td>){2,6})", re.IGNORECASE
)
_BOLTON_CELL_RE = re.compile(r"<td[^>]*>([^<]*)</td>")
_BOLTON_DISTANCE_RE = re.compile(r"([\d.]+)\s*miles")


def fetch_bolton() -> list[dict]:
    """Bolton's school directory has an individual HTML profile page
    per school (crawled from the paginated Primary and Secondary
    category listings, since the exact page count shifts as schools
    open/close) with a "last distance offered" table row giving 5
    years of figures side by side (oldest first) - the rightmost
    non-blank cell is the most recent oversubscribed year. School name
    comes from the page's own link text on the category listing page
    (already the full official name), not re-derived from the profile
    page itself.
    """
    urls: dict[str, str] = {}
    for category_url in _BOLTON_CATEGORY_URLS:
        page = 1
        while True:
            url = category_url if page == 1 else f"{category_url}/{page}"
            resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
            resp.raise_for_status()
            links = re.findall(
                r'<a class="[^"]*list__link[^"]*" href="(/directory-record/[^"]+)">([^<]+)</a>', resp.text
            )
            if not links:
                break
            for path, name in links:
                urls[path] = name.replace("&#039;", "'").replace("&amp;", "&").strip()
            if f"{category_url}/{page + 1}" not in resp.text:
                break
            page += 1

    records = []
    for path, name in urls.items():
        resp = httpx.get(f"https://www.bolton.gov.uk{path}", timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        row_match = _BOLTON_ROW_RE.search(resp.text)
        if not row_match:
            continue
        cells = _BOLTON_CELL_RE.findall(row_match.group(1))
        for cell in reversed(cells):
            distance_match = _BOLTON_DISTANCE_RE.search(cell)
            if distance_match:
                records.append({"school_name": name, "last_distance_miles": float(distance_match.group(1))})
                break
    return records


_SALFORD_PRIMARY_URL = "https://www.salford.gov.uk/media/1x4phafq/primary-schools-2025-intake.pdf"
_SALFORD_SECONDARY_URL = (
    "https://www.salford.gov.uk/schools-and-learning/schools-admissions/secondary/"
    "how-school-places-were-allocated-last-year/secondary-school-allocation-history-on-offer-day-1-march-2025/"
)
_SALFORD_DISTANCE_RE = re.compile(r"distance of ([\d.]+) miles")
_SALFORD_ROW_RE = re.compile(r"<tr>(.*?)</tr>", re.DOTALL)
_SALFORD_CELL_RE = re.compile(r"<td[^>]*>(.*?)</td>", re.DOTALL)


def fetch_salford() -> list[dict]:
    """Salford's Primary intake PDF and Secondary allocation-history
    webpage (not a PDF - the HTML page itself is a plain table) both
    describe the decisive distance in prose ("All applicants in
    categories 1-6 to a distance of 0.906 miles") rather than a
    dedicated numeric column, alongside undersubscribed rows ("All
    applicants offered places") or faith-school rows with no distance
    criterion at all ("Please contact school for finer detail") -
    searched for generically rather than parsed positionally, since
    both sources use the same prose phrasing.
    """
    records = []

    resp = httpx.get(_SALFORD_PRIMARY_URL, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or not row[0]:
                    continue
                joined = " ".join(cell for cell in row if cell)
                match = _SALFORD_DISTANCE_RE.search(joined)
                if match:
                    name = row[0].replace("\n", " ").strip()
                    records.append({"school_name": name, "last_distance_miles": float(match.group(1))})

    resp = httpx.get(_SALFORD_SECONDARY_URL, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    for row_html in _SALFORD_ROW_RE.findall(resp.text):
        cells = _SALFORD_CELL_RE.findall(row_html)
        if len(cells) < 5:
            continue
        joined = " ".join(cells)
        match = _SALFORD_DISTANCE_RE.search(joined)
        if not match:
            continue
        name = re.sub(r"<[^>]+>", "", cells[0]).strip()
        if name:
            records.append({"school_name": name, "last_distance_miles": float(match.group(1))})

    return records


_KNOWSLEY_INDEX_URLS = [
    "https://www.knowsley.gov.uk/education-and-schools/school-admissions/apply-primary-school-september-2026/oversubscribed-primary",
    "https://www.knowsley.gov.uk/education-and-schools/school-admissions/apply-secondary-school-september-2026/secondary-school",
]
_KNOWSLEY_DISTANCE_RE = re.compile(r"lived ([\d.]+)\s*(?:miles\s*)?from the school")


def fetch_knowsley() -> list[dict]:
    """Knowsley crawls two index pages (Primary/Secondary "oversubscribed
    schools") for individual per-school "oversubscription breakdown"
    PDFs, one per oversubscribed school (the set changes yearly).
    School name is the first line of the PDF's own text; the decisive
    distance is prose ("...lived 0.275 miles from the school (as the
    crow flies)") - the Primary documents sometimes omit the word
    "miles" from that sentence while Secondary always includes it, so
    it's made optional in the pattern.
    """
    records = []
    for index_url in _KNOWSLEY_INDEX_URLS:
        resp = httpx.get(index_url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        pdf_urls = set(re.findall(r'href="(https://www\.knowsley\.gov\.uk/sites/default/files/[^"]+\.pdf)"', resp.text))
        for pdf_url in pdf_urls:
            pdf_resp = httpx.get(pdf_url, timeout=30, follow_redirects=True, headers=HEADERS)
            pdf_resp.raise_for_status()
            with pdfplumber.open(io.BytesIO(pdf_resp.content)) as pdf:
                text = pdf.pages[0].extract_text() or ""
            name = text.split("\n")[0].strip()
            match = _KNOWSLEY_DISTANCE_RE.search(text)
            if name and match:
                records.append({"school_name": name, "last_distance_miles": float(match.group(1))})
    return records


def fetch_sefton() -> list[dict]:
    """Sefton Council's "Schools Admissions Information Guide" - a
    huge (200+ page) composite prospectus with a per-school profile
    section, republished each year at a new URL (find the current one
    via sefton.gov.uk's "startingschool" page). The school name is the
    first line of text on its profile's first page; "Table 2: How
    places were allocated" has an "If oversubscribed furthest distance
    (miles)" column (most other header cells on this table extract as
    reversed text, but this one doesn't, so it's used as the anchor)
    with one row per recent year, most recent first - takes that first
    row's value. Faith/academy schools that set their own admissions
    aren't included in this table and are naturally skipped.
    """
    url = "https://www.sefton.gov.uk/media/oavhszda/sefton-schools-admissions-information-guide-2026.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table or not any("furthest distance" in str(c).lower() for c in table[0]):
                continue
            if len(table) < 2 or not table[1] or not table[1][-1]:
                continue
            try:
                distance = float(table[1][-1].strip())
            except ValueError:
                continue
            text = page.extract_text() or ""
            name = text.split("\n")[0].strip()
            if name:
                records.append({"school_name": name, "last_distance_miles": distance})
    return records


def fetch_newham() -> list[dict]:
    """London Borough of Newham's "Starting school" (primary) and
    "Starting secondary school" (secondary) prospectuses - composite
    PDFs (~100-120 pages) republished each year at a new file ID
    (find current ones via newham.gov.uk's own site search for
    "starting school in newham" / "starting secondary school in
    newham" if these 404). There's no real PDF table here - the
    "How places were offered" section (a handful of pages near the
    front covering most community schools) and each individual
    voluntary-aided/academy school's own one-page admissions profile
    (scattered throughout the rest of the document) all render as
    plain text with one line per school, e.g. "Brampton Manor Academy
    460 2666 17 6 N/A 144 0 N/A 0 N/A N/A 293 0 460 All Other 0.816".
    School names can themselves contain digits (e.g. "School 21"), so
    rather than trying to isolate the name from the number of columns
    (which varies row to row), the parser anchors on the *trailing*
    "<criterion> <distance in miles>" and takes everything before the
    first digit as the name - safe because undersubscribed schools
    (ending "N/A N/A", no trailing float) are the only ones where that
    could be ambiguous, and those are skipped anyway since there's no
    distance to extract. Every real row appears twice (each double-page
    spread renders identically as two separate PDF pages) - harmless,
    de-duplicated downstream by URN.
    """
    urls = [
        "https://www.newham.gov.uk/downloads/file/9671/starting-school-in-newham-2026-",
        "https://www.newham.gov.uk/downloads/file/9673/starting-secondary-school-in-newham-full-digital-version",
    ]
    name_dist_re = re.compile(r"^([A-Za-z][A-Za-z'.\-&, ]+?)\s+\d.*\s(\d+\.\d+)\s*$")
    records = []
    for url in urls:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                for line in text.split("\n"):
                    match = name_dist_re.match(line.strip())
                    if not match:
                        continue
                    name = match.group(1).strip()
                    if name.lower() == "total":
                        continue
                    records.append({"school_name": name, "last_distance_miles": float(match.group(2))})
    return records


def fetch_bexley() -> list[dict]:
    """London Borough of Bexley's "Admission to Secondary Schools"
    booklet - a genuine PDF table (page 12 of the 2023-2024 edition)
    headed "School | 5 years ago | 4 years ago | ... | @ <date>",
    giving each selective/oversubscribed secondary school's offer
    distance for the last several years side by side; the last
    column is the most recent. Cells are either "<n> miles\\nStraight
    line"/"<n> miles\\nRoad route" (take the number, ignore the
    measurement method) or "All applicants offered places" /
    "All selective applicants offered places" for schools that
    weren't oversubscribed that year (skipped - no useful distance).
    No newer edition of this booklet with a more recent last-column
    date could be found published at a stable URL as of 2026; the
    2023-2024 edition's last column (dated July 2022) is used as-is,
    same situation as other authorities' "varies"-labelled sources.
    Bexley's primary school admissions booklet was also checked but
    contains only prose admissions policies per school, no numeric
    distance table - primary schools aren't covered here.
    """
    url = "https://www.bexley.gov.uk/sites/default/files/2023-03/Admission-to-secondary-schools-2023-2024.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    dist_re = re.compile(r"([\d.]+)\s*miles")
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table or not table[0] or table[0][0] != "School":
                continue
            for row in table[1:]:
                if not row or not row[0]:
                    continue
                name = row[0].replace("\n", " ").strip()
                last_cell = row[-1]
                if not last_cell:
                    continue
                match = dist_re.search(last_cell)
                if match:
                    records.append({"school_name": name, "last_distance_miles": float(match.group(1))})
    return records


_LIVED_RE = re.compile(r"lived\s+([\d.]+)\s+miles", re.IGNORECASE)


def fetch_havering() -> list[dict]:
    """London Borough of Havering's "Infant and Primary School
    Statistics" and "Secondary School Statistics" PDFs - the
    "downloads/download/<id>/<slug>" landing page for each just
    republishes a fresh "downloads/file/<id>/<slug>-<year>" link
    every year rather than replacing the file in place, so this
    fetches the current year's direct file link (update the numeric
    IDs below if they 404 - the landing pages list every past year's
    ID). Genuine PDF tables, but with reversed header text (same
    quirk as Sefton's table - most header cells extract backwards,
    e.g. "elbaliavA\\nsecalP" for "Places\\nAvailable" - only the data
    rows are used, not the headers). The last column is "Furthest
    Distance Allocation (in KM's)", "N/A" for schools that weren't
    oversubscribed, occasionally with trailing asterisks (e.g.
    "1.153**" - footnote markers) stripped by only taking the
    leading numeric run. The "Totals" summary row at the end of each
    table is skipped explicitly since it isn't a school.
    """
    urls = [
        "https://www.havering.gov.uk/downloads/file/7022/infant-and-primary-school-statistics-2025",
        "https://www.havering.gov.uk/downloads/file/6919/secondary-school-statistics-2025",
    ]
    dist_re = re.compile(r"^([\d.]+)")
    records = []
    for url in urls:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table:
                    continue
                for row in table:
                    if not row or not row[0]:
                        continue
                    name = row[0].replace("\n", " ").strip()
                    if name.lower() == "totals":
                        continue
                    last_cell = row[-1]
                    if not last_cell:
                        continue
                    match = dist_re.match(last_cell.strip())
                    if match:
                        km = float(match.group(1))
                        records.append({"school_name": name, "last_distance_miles": km * 1000 / _METRES_PER_MILE})
    return records


_HILLINGDON_NA_RE = re.compile(r"^N/A ?\([A-Z]\)$")


def fetch_hillingdon() -> list[dict]:
    """Hillingdon Council's main "School Admissions" brochure - one
    combined ~100+ page PDF (republished each year at a new file ID -
    find the current one via hillingdon.gov.uk's own site search for
    "School admissions <year> brochure" if this 404s) covering
    everything: admissions criteria prose, a "furthest distance
    offered" table for primary schools (3 pages) and one for
    secondary schools (1 page), THEN staff contact-details tables
    later in the document whose last column happens to be a small
    integer (published admission number) that would otherwise also
    look numeric. Only tables containing at least one cell matching
    "N/A (<letter>)" (e.g. "N/A (U)" for undersubscribed, "N/A (F)"
    for filled on other criteria) are treated as real distance
    tables - that marker is unique to the distance tables and reliably
    excludes the contact-details tables. Primary schools with two
    forms of entry (e.g. a linked nursery class) get two table rows,
    the second with a blank name cell - the school name is carried
    forward from the last non-blank cell. Distances are in metres.
    """
    url = "https://www.hillingdon.gov.uk/media/10152/School-admissions-2026-brochure/pdf/c1School_Admissions_Main_brochure_2026.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            if not any(row and row[-1] and _HILLINGDON_NA_RE.match(str(row[-1]).strip()) for row in table):
                continue
            last_name = None
            for row in table:
                if not row:
                    continue
                if row[0]:
                    last_name = row[0].replace("\n", " ").strip().rstrip("*").strip()
                if not last_name:
                    continue
                last_cell = row[-1]
                if not last_cell:
                    continue
                try:
                    metres = float(str(last_cell).strip())
                except ValueError:
                    continue
                records.append({"school_name": last_name, "last_distance_miles": metres / _METRES_PER_MILE})
    return records


_MIDDLESBROUGH_ROW_RE = re.compile(
    r"^\d{3,4}\s*(.+?)\s+\d+\s+[YN]\s+\d+\s+(?:\d+|N/A)\s+.+?\s+([\d.]+|Not Known)\s*$"
)


def fetch_middlesbrough() -> list[dict]:
    """Middlesbrough Council's combined "School admission statistics"
    PDF - one file covering several years of both primary and
    secondary intakes, most recent year first; only the first two
    pages (most recent primary, most recent secondary) are used. The
    PAGE URL slug changes periodically when a new year's edition is
    published - re-resolve it from
    middlesbrough.gov.uk/schools-and-education/school-admissions/
    school-admissions-allocation-statistics/ (a Cloudflare-protected
    page that blocks direct fetching, so search site:middlesbrough.gov.uk
    "school admissions statistics" instead) if this 404s. Primary
    pages render as plain text (not a real PDF table), one line per
    school: "<DfE number> <name> <PAN> <Y/N oversubscribed> <total
    offers> <waiting list> <last criteria text> <distance in miles or
    'Not Known'>". Unlike most sources here, Middlesbrough publishes a
    "furthest distance of the last place allocated" even for schools
    that weren't oversubscribed, which is meaningless as a cutoff (it's
    just whoever happened to be last processed) and occasionally wildly
    large (e.g. 18 miles for a small local primary) - distances over 10
    miles are dropped as implausible for this compact urban borough,
    same rationale as Warwickshire's plausibility cap.
    """
    url = "https://www.middlesbrough.gov.uk/media/vuxofasc/school-admissions-statistics-2022-2026.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages[:2]:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                match = _MIDDLESBROUGH_ROW_RE.match(line.strip())
                if not match:
                    continue
                name, dist = match.group(1).strip(), match.group(2)
                if dist == "Not Known":
                    continue
                distance = float(dist)
                if distance > 10:
                    continue
                records.append({"school_name": name, "last_distance_miles": distance})
    return records


_HARTLEPOOL_DIST_RE = re.compile(r"^([\d.]+)\s*km$", re.IGNORECASE)


def fetch_hartlepool() -> list[dict]:
    """Hartlepool Borough Council's "Admission to Secondary School
    Allocation Report" PDF - a genuine table (reversed header text,
    same quirk as Sefton/Havering) with one 3-row block per school:
    the first row already carries the most recent year's ("2025")
    data including the "Last Distance Offered" (km) as the last
    cell, with the prior year's block following immediately after -
    only the first (most recent) block per school is used, so no
    year-column parsing is needed, just "first row with a name". "n/a"
    in the distance cell means the school wasn't oversubscribed that
    year and is skipped. No equivalent primary-schools allocation
    report with per-school distance data could be found published at
    a stable URL - only secondary schools are covered here.
    """
    url = "https://www.hartlepool.gov.uk/downloads/file/1216/admission-to-secondary-school-allocation-report-2024-and-2025"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue
            for row in table:
                if not row or not row[0]:
                    continue
                name = row[0].replace("\n", " ").strip()
                if name == "School":
                    continue
                last_cell = row[-1]
                if not last_cell:
                    continue
                match = _HARTLEPOOL_DIST_RE.match(str(last_cell).strip())
                if not match:
                    continue
                km = float(match.group(1))
                records.append({"school_name": name, "last_distance_miles": km * 1000 / _METRES_PER_MILE})
    return records


_SOUTHEND_DIST_RE = re.compile(r"distance of\s+([\d.]+)\s*mile", re.IGNORECASE)


def fetch_southend() -> list[dict]:
    """Southend-on-Sea City Council's primary admissions booklet - a
    composite prospectus with one per-school profile page each,
    republished each year at a new file ID (find the current one via
    southend.gov.uk's own site search for "primary admission booklet"
    if this 404s). Each profile page's first line is the school name;
    somewhere in its "last child admitted" paragraph is a sentence
    like "...the last child was admitted under admission criterion
    catchment area at a distance of 0.245 miles from the school" (the
    exact wording around it varies - "at a distance of" vs "with a
    distance of" - and it sometimes wraps across two lines, so the
    page text has its whitespace collapsed before a single regex
    anchored on "distance of <n> mile(s)" is applied). Most Southend
    primary schools use catchment areas rather than distance and have
    no such sentence at all (naturally skipped - no match). The
    equivalent secondary admissions booklet was checked too but has no
    per-school distance data anywhere - secondary schools here mostly
    use defined catchment areas or aptitude, not distance - so only
    primary schools are covered.
    """
    url = "https://www.southend.gov.uk/downloads/file/9043/primary-admission-booklet-2026-27"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines = text.split("\n")
            if not lines or not lines[0].strip():
                continue
            blob = re.sub(r"\s+", " ", text)
            match = _SOUTHEND_DIST_RE.search(blob)
            if match:
                records.append({"school_name": lines[0].strip(), "last_distance_miles": float(match.group(1))})
    return records


_BRACKNELL_NAME_RE = re.compile(r"Offer Day:.*?\d{4}\s+(.+?)\s*\n\s*Preferences Received", re.S)
_BRACKNELL_DIST_RE = re.compile(r"to a distance of\s+([\d.]+)\s*miles", re.IGNORECASE)


def fetch_bracknell_forest() -> list[dict]:
    """Bracknell Forest Council's "How school places are allocated"
    page links to one "allocation breakdown" PDF per individual
    school (primary and secondary), crawled dynamically since the
    URLs change every year. Two of the linked PDFs ("...community-
    schools-allocation-breakdown-..." and "...voluntary-schools-
    allocation-breakdown-...") cover MULTIPLE community/voluntary
    primary schools in one file with a two-column layout (school
    name/PAN in one column, prose explanation in the other) that
    pdfplumber's text extraction interleaves unpredictably between
    columns - real risk of a distance figure being attributed to the
    wrong neighbouring school - so those two are deliberately
    excluded; only genuinely single-school PDFs are used. Each of
    those has the school's name appearing consistently between
    "National Offer Day: <date>" and "Preferences Received" (more
    reliable than the page's opening lines, whose order relative to
    the "Allocation breakdown for admission to..." header varies
    between documents), and - only for oversubscribed schools - a
    sentence ending "...to a distance of X miles from the school".
    Schools that weren't oversubscribed have no such sentence and
    are naturally skipped.
    """
    index_url = "https://www.bracknell-forest.gov.uk/schools-and-learning/schools/school-admissions/how-school-places-are-allocated"
    print(f"  Downloading index {index_url}")
    resp = httpx.get(index_url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    pdf_urls = re.findall(r'href="(https://www\.bracknell-forest\.gov\.uk/sites/default/files/[^"]+?-allocation-breakdown-[^"]+?\.pdf)"', resp.text)
    pdf_urls = [
        u for u in dict.fromkeys(pdf_urls)
        if "community-schools-allocation-breakdown" not in u and "voluntary-schools-allocation-breakdown" not in u
    ]

    records = []
    for url in pdf_urls:
        print(f"  Downloading {url}")
        r = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        if r.status_code != 200:
            continue
        with pdfplumber.open(io.BytesIO(r.content)) as pdf:
            text = pdf.pages[0].extract_text() or ""
        name_match = _BRACKNELL_NAME_RE.search(text)
        dist_match = _BRACKNELL_DIST_RE.search(text)
        if name_match and dist_match:
            records.append({"school_name": name_match.group(1).strip(), "last_distance_miles": float(dist_match.group(1))})
    return records


def fetch_islington() -> list[dict]:
    """Islington Council's "Cut-off distance for schools" page - a
    genuine HTML table (two of them: primary, then secondary) with
    one row per school and three year columns side by side
    ("Distance (miles) in 2026-27" / "2025-26" / "2024-25", most
    recent first); takes the first column with a real number rather
    than "Not applicable" (undersubscribed that year). Banded
    secondary schools (e.g. Central Foundation Boys' School, which
    publishes a separate cut-off per aptitude/distance band) appear
    as multiple rows with " Band 1"/"Band 2" etc suffixed onto the
    school name - all fuzzy-match to the same real school and
    de-duplicate downstream to one URN, same as a school appearing
    twice for any other reason. <thead> is stripped before parsing
    so its header cells (which also use <th>) aren't mistaken for a
    school row.
    """
    import html as html_module

    url = (
        "https://www.islington.gov.uk/children-and-families/schools/"
        "apply-for-a-school-place/school-admissions-information/cut-off-distance-maps"
    )
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    text = re.sub(r"<thead>.*?</thead>", "", resp.text, flags=re.S)

    row_re = re.compile(
        r"<tr>\s*<th>(.*?)</th>\s*<td>.*?</td>\s*<td>(.*?)</td>\s*<td>(.*?)</td>\s*<td>(.*?)</td>\s*</tr>", re.S
    )
    records = []
    for match in row_re.finditer(text):
        name = html_module.unescape(re.sub(r"<.*?>", "", match.group(1))).strip()
        for cell in (match.group(2), match.group(3), match.group(4)):
            cell_clean = html_module.unescape(re.sub(r"&nbsp;|<.*?>", "", cell)).strip()
            try:
                distance = float(cell_clean)
            except ValueError:
                continue
            records.append({"school_name": name, "last_distance_miles": distance})
            break
    return records


def fetch_southampton() -> list[dict]:
    """Southampton City Council publishes two separate "how places
    were offered at oversubscribed schools" documents: a PDF for
    secondary schools with three years' tables stacked on separate
    pages (most recent year last - only that page is used) giving a
    "Distance of last child admitted" as the final column of each
    school's block (column position varies row to row since each
    school's admissions criteria differ, so the last float-typed
    cell in the row is taken rather than a fixed index); and an XLSX
    for infant/primary schools with one sheet per year (most recent
    sheet used, named by year), where "Community"/"Voluntary
    Controlled"/"Trust" schools appear in genuine numeric-column
    tables (again: last float-typed cell = the distance) but
    "Academies"/"Voluntary Aided" schools instead get a prose
    "Criterion" cell reading "...This child lived 1.761 miles from
    the school." (regex-extracted). Rows with neither a trailing
    float nor a "lived ... miles" phrase (headers, section titles,
    explanatory notes, schools that weren't oversubscribed) are
    naturally skipped since no distance can be found. Both source
    files are hosted at fairly stable per-year media IDs.
    """
    records = []

    url = "https://www.southampton.gov.uk/media/ur2kqqix/how-places-were-offered-at-oversubscribed-schools-last-3-years.pdf"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        page = pdf.pages[-1]
        table = page.extract_table()
        for row in table or []:
            if not row or not row[0]:
                continue
            name = row[0].strip()
            distance = None
            for cell in reversed(row[1:]):
                if cell is None:
                    continue
                try:
                    distance = float(cell)
                    break
                except ValueError:
                    continue
            if distance is not None:
                records.append({"school_name": name, "last_distance_miles": distance})

    url2 = (
        "https://www.southampton.gov.uk/media/jwkpgyaz/"
        "how-places-were-offered-at-oversubscribed-infant-and-primary-schools-for-year-r-entry-september-2024-2023-and-2022.xlsx"
    )
    print(f"  Downloading {url2}")
    resp2 = httpx.get(url2, timeout=60, follow_redirects=True, headers=HEADERS)
    resp2.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp2.content), data_only=True)
    latest_sheet = max(wb.sheetnames, key=lambda s: int(s))
    for row in wb[latest_sheet].iter_rows(values_only=True):
        if not row or not row[0] or not isinstance(row[0], str) or row[0] == "School":
            continue
        name = row[0].strip()
        distance = None
        for cell in reversed(row[1:]):
            if isinstance(cell, float):
                distance = cell
                break
        if distance is None:
            for cell in row[1:]:
                if isinstance(cell, str):
                    match = _LIVED_RE.search(cell)
                    if match:
                        distance = float(match.group(1))
                        break
        if distance is not None:
            records.append({"school_name": name, "last_distance_miles": distance})

    return records


_BRIGHTON_HOVE_HEADER_RE = re.compile(r"([A-Z][A-Za-z'&,.\- ]{2,60}?(?:School)) - \d+ places")
_BRIGHTON_HOVE_DISTANCE_RE = re.compile(r"furthest child offered a place in priority \d+ lives ([\d.]+) metres")


def fetch_brighton_and_hove() -> list[dict]:
    """Brighton & Hove City Council publishes its Reception allocation
    outcomes as prose HTML (not a table) - one paragraph per
    oversubscribed school reading "<School> - NN places ... The
    furthest child offered a place in priority 5 lives NNN.NN metres
    from the school." Only schools that were actually oversubscribed
    down to the "other children" priority carry this sentence, so most
    schools on the page are correctly skipped rather than guessed at.
    Distances are in metres, converted to miles here.
    """
    url = (
        "https://www.brighton-hove.gov.uk/schools-and-learning/school-policies-reports-strategies-and-other-documents/"
        "allocation-infantprimary-school-reception-places-september-2025"
    )
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    text = re.sub(r"<[^>]+>", " ", resp.text)
    text = text.replace("&amp;", "&").replace("&rsquo;", "'").replace("&lsquo;", "'").replace("&ndash;", "-")
    text = re.sub(r"\s+", " ", text)

    records = []
    headers = list(_BRIGHTON_HOVE_HEADER_RE.finditer(text))
    for i, m in enumerate(headers):
        name = m.group(1).strip()
        chunk_end = headers[i + 1].start() if i + 1 < len(headers) else len(text)
        chunk = text[m.end():chunk_end]
        dm = _BRIGHTON_HOVE_DISTANCE_RE.search(chunk)
        if dm:
            records.append({"school_name": name, "last_distance_miles": float(dm.group(1)) / _METRES_PER_MILE})
    return records


_BROMLEY_PRIMARY_URL = "https://www.bromley.gov.uk/downloads/file/28/primary-education-in-bromley-2025-2026"
_BROMLEY_SECONDARY_URL = "https://www.bromley.gov.uk/downloads/file/1863/secondary-education-in-bromley-2024-2025"
_BROMLEY_TOKEN = r"(?:N/A|Church\s+Criteria|[\d.]+)"
_BROMLEY_ROWTAIL_RE = re.compile(
    rf"(?:Information\s+not\s+available|\d{{1,3}}\s+\d{{1,3}})\s+({_BROMLEY_TOKEN})\s+({_BROMLEY_TOKEN})\s+({_BROMLEY_TOKEN})"
)
_BROMLEY_FOOTER_RE = re.compile(
    r"\*?\s*The distance information provided is as at the relevant national offer day\s*"
    r"|N/A - school undersubscribed LA Allocations made\s*"
)
_BROMLEY_SECONDARY_ROW_RE = re.compile(
    r"^([A-Za-z][A-Za-z0-9'&,.\- ]+?) (?:\d+\s+\d+|N/A\s+N/A|Information not available) "
    r"([\d.]+|N/A) ([\d.]+|N/A) ([\d.]+|N/A)$"
)


def _bromley_first_distance(tokens) -> float | None:
    for token in tokens:
        if re.fullmatch(r"[\d.]+", token):
            return float(token)
    return None


def fetch_bromley() -> list[dict]:
    """Bromley Council's "Primary/Secondary Education in Bromley"
    prospectus booklets each include a "Distances and appeals" table
    (School / Heard / Upheld / Distance in miles for the last 3 years).
    Already in miles; takes the first (most recent) available year
    per school, same convention as Haringey.

    The primary/junior table is prose-wrapped (not a real PDF table)
    and a handful of faith schools show "Church Criteria" instead of a
    distance for one or more years, which breaks the normal
    row-boundary regex and would otherwise merge that school's name
    with the next school's - two real schools' data smeared into one
    bogus combined-name record. Any chunk that ends up looking like
    more than one school (too long, or more than one "School"/
    "Academy" keyword) is dropped rather than risk that bleed.

    Secondary schools that allocate by banding or test score (grammar/
    partially-selective schools) have no single "last distance
    offered" - explicitly skipped rather than picking one band's
    figure and presenting it as the school's cutoff.
    """
    records = []

    print(f"  Downloading {_BROMLEY_PRIMARY_URL}")
    resp = httpx.get(_BROMLEY_PRIMARY_URL, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        table_text = ""
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if "Heard" in page_text and "Upheld" in page_text and "Distance in miles" in page_text:
                table_text += page_text + " "

    text = re.sub(r"\s+", " ", table_text)
    start = text.find("Alexandra Infant School")
    matches = list(_BROMLEY_ROWTAIL_RE.finditer(text)) if start >= 0 else []
    prev_end = start
    for m in matches:
        if m.start() < start:
            continue
        name_chunk = text[prev_end:m.start()].strip()
        name_chunk = _BROMLEY_FOOTER_RE.sub(" ", name_chunk).strip()
        name_chunk = re.split(r"2025\s+2024\s+2023", name_chunk)[-1].strip()
        name_chunk = re.sub(r"^\d{1,3}\s*", "", name_chunk)
        name_chunk = re.sub(r"\s+", " ", name_chunk).strip()
        prev_end = m.end()

        keyword_count = len(re.findall(r"\b(School|Academy)\b", name_chunk))
        if not name_chunk or len(name_chunk) > 55 or keyword_count > 1:
            continue
        distance = _bromley_first_distance(m.groups())
        if distance is not None:
            records.append({"school_name": name_chunk, "last_distance_miles": distance})

    print(f"  Downloading {_BROMLEY_SECONDARY_URL}")
    resp = httpx.get(_BROMLEY_SECONDARY_URL, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if not ("Heard" in page_text and "Upheld" in page_text and "Distance in miles" in page_text):
                continue
            for line in page_text.split("\n"):
                line = line.strip()
                if not line or "banding" in line.lower() or "test score" in line.lower():
                    continue
                m = _BROMLEY_SECONDARY_ROW_RE.match(line)
                if not m:
                    continue
                distance = _bromley_first_distance(m.groups()[1:])
                if distance is not None:
                    records.append({"school_name": m.group(1).strip(), "last_distance_miles": distance})

    return records


_CAMDEN_SECONDARY_URL = (
    "https://www.stlukesschool.org.uk/wp-content/uploads/2025/03/"
    "Secondary-schools-in-Camden-admissions-guide-2025-WEB.pdf"
)
_CAMDEN_ROW_RE = re.compile(r"^([A-Za-z][A-Za-z'&,.\- ]+?) (n/a|[\d.]+) (n/a|[\d.]+)$", re.IGNORECASE)


def fetch_camden() -> list[dict]:
    """Camden's own site (camden.gov.uk) is blocked from this
    environment, but its "Secondary Schools in Camden" admissions
    guide - which includes a "cut off distances" table for the
    previous two years - is also mirrored as a PDF on individual
    Camden secondary schools' own sites (found here via St Luke's
    Church of England School). No equivalent primary-school distance
    table was locatable, so this is secondary-only, same situation as
    Hartlepool. Already in miles; takes 2024 if present else 2023.
    Camden School for Girls allocates by 4 separate bands rather than
    a single distance and is correctly skipped (its row doesn't match
    the plain "name value value" pattern).
    """
    print(f"  Downloading {_CAMDEN_SECONDARY_URL}")
    resp = httpx.get(_CAMDEN_SECONDARY_URL, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if "cut off distances" not in page_text:
                continue
            for line in page_text.split("\n"):
                line = line.strip()
                if not line or "Band" in line or line.lower().startswith("school "):
                    continue
                m = _CAMDEN_ROW_RE.match(line)
                if not m:
                    continue
                distance = _bromley_first_distance(m.groups()[1:])
                if distance is not None:
                    records.append({"school_name": m.group(1).strip(), "last_distance_miles": distance})
    return records


_WESTMINSTER_URL = (
    "https://www.westminster.gov.uk/sites/default/files/media/documents/"
    "WCC%20-%20Primary%20Admissions%20Brochure%202025.pdf"
)
_WESTMINSTER_NAME_RE = re.compile(r"retsnimtseW\s+(.*?)\s+(?:SCHOOL INFORMATION|SUMMARISED ADMISSION CRITERIA)")
_WESTMINSTER_DIST_RE = re.compile(r"up to[^.]{0,40}?(\d+\.\d+)[^.]{0,25}?miles?", re.IGNORECASE)


def fetch_westminster() -> list[dict]:
    """Westminster's primary admissions brochure profiles two schools
    per page, side by side, each with its own "HOW PLACES WERE OFFERED
    IN 2024" summary. Critically, pdfplumber's default text extraction
    reads across the whole page rather than down each column, which
    would interleave the two schools' text and risk attributing one
    school's distance to its neighbour - so each page is split into a
    left-half and right-half crop first (using page.width/2) and each
    half is read independently, keeping every school's text with only
    that school.

    Even within one half there's a second, smaller two-column split
    (criteria text vs. an address/phone/email sidebar) that can still
    drop an odd word into the middle of a sentence (e.g. "...distance
    Email of 0.505 of a mile..."), so the distance regex tolerates a
    short run of intervening text between "up to" and the figure
    rather than requiring an exact phrase match - the page consistently
    uses "up to ... 0.NNN ... mile(s)" with only one number in that
    span, so this stays a single unambiguous figure per school rather
    than a guess. Most schools here are small enough that they never
    reach the distance criterion at all (admitted on faith/sibling
    priorities alone) and are correctly left without a figure.

    The reversed "...retsnimtseW" text preceding each school's name is
    sideways watermark text (page furniture, not a heading);
    "retsnimtseW" ("Westminster" reversed) is used as the anchor for
    where each school's real name begins.
    """
    print(f"  Downloading {_WESTMINSTER_URL}")
    resp = httpx.get(_WESTMINSTER_URL, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if "SCHOOL INFORMATION" not in page_text:
                continue
            width = page.width
            for half in (page.crop((0, 0, width / 2, page.height)), page.crop((width / 2, 0, width, page.height))):
                half_text = re.sub(r"\s+", " ", half.extract_text() or "")
                name_match = _WESTMINSTER_NAME_RE.search(half_text)
                dist_match = _WESTMINSTER_DIST_RE.search(half_text)
                if name_match and dist_match:
                    name = name_match.group(1).strip().title()
                    records.append({"school_name": name, "last_distance_miles": float(dist_match.group(1))})
    return records


_RBKC_URL = (
    "https://www.rbkc.gov.uk/sites/default/files/media/documents/RBKC%20-%20Primary%20Admissions%20Brochure%202026.pdf"
)
_RBKC_NAME_RE = re.compile(r"School information\s+(.*?)\s+Summarised admi")
_RBKC_BLOCK_RE = re.compile(r"How places were offered(.*?)(?:Appeals|Footnote)")
_RBKC_DIST_RE = re.compile(r"\((?:up to (?:a distance of )?|distance )(\d+\.\d+)[^)]*mile[^)]*\)", re.IGNORECASE)


def fetch_kensington_and_chelsea() -> list[dict]:
    """Royal Borough of Kensington and Chelsea's primary admissions
    brochure (same shared-service format as neighbouring Westminster's,
    but one school per page rather than two) has a "How places were
    offered" box per school listing each oversubscription category in
    priority order, e.g. faith schools often show both a "Foundation"
    (faith-priority) figure and an "Open" (general distance) figure,
    such as "Foundation: 12 pupils (distance 1.217 miles) ... Open: 5
    pupils (up to a distance of 0.386 miles)". Earlier categories in
    that list are consistently followed by "(all ... applicants
    offered)" rather than a distance UNLESS that category itself is
    where the school's places ran out, so the category that actually
    determines "last child offered a place" overall is always the
    LAST distance figure mentioned in the box - taking the first one
    instead would sometimes report a faith-only cutoff as if it were
    the general distance cutoff. A small in-page column layout also
    interleaves unrelated sentence fragments into the middle of this
    box; whitespace is collapsed to keep the regex working across that
    noise the same way it does for Westminster.
    """
    print(f"  Downloading {_RBKC_URL}")
    resp = httpx.get(_RBKC_URL, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if "School information" not in page_text or "Summarised admi" not in page_text:
                continue
            text = re.sub(r"\s+", " ", page_text)
            name_match = _RBKC_NAME_RE.search(text)
            block_match = _RBKC_BLOCK_RE.search(text)
            if not (name_match and block_match):
                continue
            distances = _RBKC_DIST_RE.findall(block_match.group(1))
            if distances:
                records.append({"school_name": name_match.group(1).strip(), "last_distance_miles": float(distances[-1])})
    return records


_TOWER_HAMLETS_URL = "https://stebon.org.uk/wp-content/uploads/2025/01/Primary-School-prospectus.pdf"
_TOWER_HAMLETS_ROW_RE = re.compile(r"^([A-Za-z][A-Za-z0-9'&,.\- ]*?) ((?:\d+ ){2,}\d+)$")


def fetch_tower_hamlets() -> list[dict]:
    """Tower Hamlets Council's own site wasn't checked directly - its
    "Starting Primary School" prospectus (which only covers community
    schools, the only ones for which the Council itself is the
    admission authority) is also mirrored on individual schools' own
    sites (found here via Stebon Primary School). Its "Summary of last
    year's application and offers" table is watermark-rotated (each
    column heading is reversed character-by-character, e.g. "loohcS"
    for "School") but the data rows themselves read normally: name
    followed by 12 numbers, of which the second-to-last is "Offered
    distance (metres)" and the last is the total number of places -
    confirmed by the last number consistently matching the school's
    published admission number (PAN). Converted from metres to miles.

    The table lists schools by a short, sometimes-abbreviated name
    (e.g. "Bygrove" rather than "Bygrove Primary School") that's too
    short relative to the full GIAS name to pass the fuzzy-match
    cutoff on its own - "Primary School" is appended to any name that
    doesn't already contain "School" before matching. One row
    ("Bonner Bethnal Green") doesn't correspond to any real Tower
    Hamlets school under that name (the actual school is "Bonner
    Primary School") and is correctly left unmatched rather than
    guessed at.
    """
    print(f"  Downloading {_TOWER_HAMLETS_URL}")
    resp = httpx.get(_TOWER_HAMLETS_URL, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    records = []
    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ""
            if "Summary of last" not in page_text or "dereffO" not in page_text:
                continue
            for line in page_text.split("\n"):
                m = _TOWER_HAMLETS_ROW_RE.match(line.strip())
                if not m:
                    continue
                name = m.group(1).strip()
                if "school" not in name.lower():
                    name += " Primary School"
                numbers = m.group(2).split()
                distance_metres = int(numbers[-2])
                records.append({"school_name": name, "last_distance_miles": distance_metres / _METRES_PER_MILE})
    return records


_WIRRAL_URL = "https://www.wirral.gov.uk/files/primary-policy-booklet2026-2027-final.pdf/download"
_WIRRAL_DIST_RE = re.compile(r"Last F2 allocated in 2025:\s*([^\n]*)")
_WIRRAL_NAME_RE = re.compile(r"[A-Z][A-Za-z'’&.\- ]+?(?:School|Academy)")


def fetch_wirral() -> list[dict]:
    """Wirral Council's primary admissions policy booklet opens with a
    directory of every primary school, each entry ending "Last F2
    allocated in 2025: <result>" - either a real category and distance
    ("Out of Zone (Category 5) - 0.723 miles"), "All on-time applicants
    offered places" (undersubscribed, correctly skipped), or an
    applicant/place count for selective schools (no single distance,
    correctly skipped). Entries are split on the "Admission Policy:
    Page NN" footer that ends every entry, then the school name is
    recovered as the last run of capitalised words ending "School" or
    "Academy" before that entry's distance sentence - taking the last
    such run (rather than the first) matters at a page break, where
    the page header/legend text can get prepended to the next entry's
    name by the text extraction.
    """
    print(f"  Downloading {_WIRRAL_URL}")
    resp = httpx.get(_WIRRAL_URL, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()

    with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    records = []
    for block in re.split(r"Admission Policy: Page \d+", full_text):
        dist_match = _WIRRAL_DIST_RE.search(block)
        if not dist_match:
            continue
        distance_match = re.search(r"([\d.]+)\s*miles", dist_match.group(1))
        if not distance_match:
            continue
        pre_text = block[:dist_match.start()].strip()
        pre_text = re.sub(r"^\d+-\d+\s+\d+\s+\d+\s+(?:Yes|No)\s*", "", pre_text)
        pre_text = re.sub(r"(?<![A-Za-z])(?:FAC|VC|VA|AC|C)(?![A-Za-z])", " ", pre_text)
        pre_text = re.sub(r"\s+", " ", pre_text).strip()
        name_matches = _WIRRAL_NAME_RE.findall(pre_text)
        name = name_matches[-1].strip() if name_matches else pre_text
        name = re.sub(r"^(?:Yes|No)\s+", "", name)
        records.append({"school_name": name, "last_distance_miles": float(distance_match.group(1))})
    return records


_BEDFORD_URLS = [
    "https://www.bedford.gov.uk/files/appendix-starting-school-2026-late-round-version.pdf",
    "https://www.bedford.gov.uk/files/appendix-secondary-2026-late-round.pdf",
]
_BEDFORD_DISTANCE_RE = re.compile(r"([\d.]+)\s*m\b")


def fetch_bedford() -> list[dict]:
    """Bedford Borough Council's "Starting School" (lower/primary) and
    "Transfer to Secondary School" (Year 7) allocation appendices -
    both clean pdfplumber tables with a final "Last place offered,
    criteria and distance" column already in metres (e.g. "Catchment,
    3777.25m" or the unicode-quoted "'any other' 1367.52m").
    Undersubscribed schools read "All offered" with no number and are
    correctly skipped rather than guessed.
    """
    records = []
    for url in _BEDFORD_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                for table in page.extract_tables():
                    for row in table:
                        if not row or not row[0] or row[0].strip() == "School":
                            continue
                        cell = row[-1]
                        if not cell:
                            continue
                        match = _BEDFORD_DISTANCE_RE.search(cell.replace("\n", " "))
                        if not match:
                            continue
                        name = row[0].replace("\n", " ").strip()
                        records.append({"school_name": name, "last_distance_miles": float(match.group(1)) / _METRES_PER_MILE})
    return records


_CENTRAL_BEDS_URLS = [
    # Starting School (lower/primary), Transfer to Middle, Transfer to Upper -
    # each year's "on-time" (national offer day) allocation info PDF, hosted
    # on Central Bedfordshire's SharePoint (linked from
    # centralbedfordshire.gov.uk's "places offered" pages for each phase -
    # follow the "On-time offers <year>" link there if these rot).
    "https://centralbedfordshirecouncil.sharepoint.com/:b:/s/Communications/IQBlmnDltkSeS7o6frAbXLJKASKVyn6jVaifuakjiHLI0gQ?e=iJKMLd&download=1",
    "https://centralbedfordshirecouncil.sharepoint.com/:b:/s/Communications/IQDv5MX6FJOCSKBdQVjP0zV3ARQPdxpJewGKuEhZhrN_gOQ?e=HX4tLo&download=1",
    "https://centralbedfordshirecouncil.sharepoint.com/:b:/s/Communications/IQAnVfWdDeEeR7GWklYbHDI6ATUDkMdR8aw48LDh-7Dr6dA?e=yjc7MP&download=1",
]
# Each oversubscribed school gets its own breakdown page headed "<Name> -
# <N> places available" (or "<Name> - places available <N>", order varies) -
# the separator glyph pdfplumber extracts is a mundane replacement character
# rather than a real "-", so matched generically as a single non-space char.
_CBC_HEADING_RE = re.compile(r"^([A-Z][^\n]{1,70}?)\s+\S\s+(?:[\d,]+\s+)?[Pp]laces available", re.MULTILINE)
_CBC_DISTANCE_RE = re.compile(r"distance of ([\d,]+\.?\d*)\s*metres")


def fetch_central_bedfordshire() -> list[dict]:
    """Central Bedfordshire Council's three-tier (lower/middle/upper)
    allocation info PDFs. Page 1 is a summary table (no per-school
    distance); subsequent pages give each oversubscribed school its
    own prose breakdown ending "...lived at a distance of NNNN.NN
    metres from the school." - undersubscribed schools get no
    breakdown page at all and are correctly absent rather than
    guessed. Schools genuinely ambiguous under fuzzy matching (e.g.
    two "Greenleas" or two "St Andrew's" sites sharing a base name)
    come out below the match cutoff and are correctly dropped by the
    shared matcher rather than handled specially here.
    """
    records = []
    for url in _CENTRAL_BEDS_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=30, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
        headings = list(_CBC_HEADING_RE.finditer(full_text))
        for i, m in enumerate(headings):
            start = m.end()
            end = headings[i + 1].start() if i + 1 < len(headings) else len(full_text)
            dist_match = _CBC_DISTANCE_RE.search(full_text[start:end])
            if not dist_match:
                continue
            metres = float(dist_match.group(1).replace(",", ""))
            records.append({"school_name": m.group(1).strip(), "last_distance_miles": metres / _METRES_PER_MILE})
    return records


_WEST_NORTHANTS_URLS = [
    # Northampton, Daventry & South, and Junior "allocations" PDFs -
    # republished each year at new /media/<id>/download IDs (find the
    # current ones via westnorthants.gov.uk's "Primary school place
    # offers" page - it's a Next.js app whose main HTML doesn't embed
    # the document links directly, so re-locate them by view-source or
    # site search rather than assuming the page's raw HTML lists them).
    "https://cms.westnorthants.gov.uk/media/29929/download",
    "https://cms.westnorthants.gov.uk/media/29930/download",
    "https://cms.westnorthants.gov.uk/media/29931/download",
]
# Only the first ("last pupil ... criterion lives X miles from the
# school") match per cell is taken - later "May/June/July round of
# reallocations" paragraphs in the same cell describe *different*,
# later distances (sometimes "from their nearest alternative school",
# a different figure entirely) for places that became free after
# National Offer Day, not the on-time figure this registry wants.
_WEST_NORTHANTS_DISTANCE_RE = re.compile(r"last pupil.*?criterion lives\s*([\d.]+)\s*miles", re.IGNORECASE | re.DOTALL)


def fetch_west_northamptonshire() -> list[dict]:
    """West Northamptonshire Council's "how places were allocated"
    PDFs (Northampton town, Daventry & South districts, and a separate
    Junior-school document) - one clean table per document with
    columns School Name / How places were allocated / Places
    remaining?, one row per school (multi-line names like "Duston
    Eldean Primary\\nSchool" wrap consistently mid-name rather than
    splitting the identifying part across rows, unlike the
    town-name-prefix wrapping that made Stoke-on-Trent and Shropshire's
    similarly-formatted booklet unsafe - see COUNCIL_COVERAGE_LOG.md).
    The distance is embedded in prose within the second column rather
    than its own column. Secondary allocations are published in a
    separate, messier document (rotated header cells, cross-page
    continuation rows with a blank school-name cell, and "linked area"
    reallocation distances that mean something different) - not
    attempted here, primary/junior only.
    """
    records = []
    for url in _WEST_NORTHANTS_URLS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=45, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if not table:
                    continue
                for row in table:
                    if not row or len(row) < 2 or not row[0] or row[0].strip() == "School Name":
                        continue
                    name = row[0].replace("\n", " ").strip()
                    cell = (row[1] or "").replace("\n", " ")
                    match = _WEST_NORTHANTS_DISTANCE_RE.search(cell)
                    if match:
                        records.append({"school_name": name, "last_distance_miles": float(match.group(1))})
    return records


# (local authority - must exactly match SchoolDetail.local_authority,
#  academic year label, fetch function)
def fetch_essex() -> list[dict]:
    """Essex County Council: "Last Distance offered, Reception and Year 2
    to 3 - Main Round Admissions 2025", released by the council as an
    Excel attachment to FOI ECC19026611 (10 Sep 2025) on its own FOI
    publication site, so a council-published figure rather than a
    third party's. Straight-line miles, furthest distance under the
    lowest criterion offered, as at national offer day. A blank
    distance means the school was undersubscribed and no distance was
    recorded, so those rows are skipped. Primary only (Reception and
    junior transfer); Essex's secondary figures were not in the
    release. Columns: DfE No., School name, Distance.
    """
    url = "https://secureapps.essex.gov.uk/Freedom_of_information/view_doc.aspx?DocID=56825"
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)
    records = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [c for c in row if c is not None]
            if len(cells) < 3:
                continue
            dfe, name, dist = cells[0], cells[1], cells[2]
            if str(name).strip().lower() == "school name":
                continue
            try:
                miles = float(str(dist).strip())
            except (TypeError, ValueError):
                continue
            if miles <= 0:
                continue
            records.append({"school_name": str(name).strip(), "last_distance_miles": miles})
    return records



def fetch_sheffield() -> list[dict]:
    """Sheffield City Council publishes two PDFs of oversubscribed
    schools for the September 2025 intake, each a single table with
    the allocations by criterion and a "distance of last place
    offered" (secondary) or "distance of last child allocated as at
    16th April" (reception) column in straight-line miles. Only
    oversubscribed schools appear, so a school missing from the file
    took everyone who applied. Names are the council's short forms
    ("Meadowhead", "Anns Grove") and are fuzzy-matched within
    Sheffield. Secondary rows carry a DfE number in column 0; the
    reception table does not, and its second page repeats no header.
    """
    sources = [
        ("https://www.sheffield.gov.uk/sites/default/files/2025-02/oversubscribed_secondary_schools_2025.pdf", 1, 11, "secondary"),
        ("https://www.sheffield.gov.uk/sites/default/files/2025-07/oversubscribed_infant_and_primary_schools.pdf", 0, 9, "reception"),
    ]
    records = []
    for url, name_col, dist_col, phase_hint in sources:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                for table in page.extract_tables():
                    for row in table:
                        if len(row) <= dist_col:
                            continue
                        name = (row[name_col] or "").replace(chr(10), " ").strip()
                        if not name or name.lower() == "school":
                            continue
                        try:
                            miles = float((row[dist_col] or "").strip())
                        except ValueError:
                            continue
                        if miles <= 0:
                            continue
                        records.append({"school_name": name, "phase_hint": phase_hint, "last_distance_miles": miles})
    return records


# Manchester's web application firewall answers every scripted request
# with 403, so the live parse below cannot be exercised from here. These
# are the figures on the page as read in a browser on 2 Sep 2026, kept as
# the fallback so a re-run still loads Manchester when the fetch is
# refused. Re-read the page when the next offer-day figures appear
# (early March) and update both the list and the date.
_MANCHESTER_FALLBACK = [
    ("Burnage Academy for Boys", 2.152),
    ("Chorlton High School", 1.375),
    ("Co-op Academy Belle Vue", 0.853),
    ("Co-op Academy North Manchester", 1.267),
    ("Didsbury High School", 0.999),
    ("Levenshulme High School", 1.623),
    ("Parrs Wood High School", 2.275),
]


def fetch_manchester() -> list[dict]:
    """Manchester City Council: "The demand for secondary school places",
    a web page listing every secondary school with its places, offers and,
    for the schools that filled on distance, the sentence "If they were in
    category N they had to live within X miles". Year 7, September 2026
    intake, offers to on-time and late applicants received by 22 Feb 2026.
    Straight-line miles to the school's centre, per the council's rules.
    Schools that took everyone, or that use their own rules and publish no
    distance, are not given a figure.
    """
    url = "https://www.manchester.gov.uk/schools-education-and-childcare/school-admissions/the-demand-for-secondary-school-places"
    print(f"  Downloading {url}")
    try:
        resp = httpx.get(url, timeout=60, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
    except httpx.HTTPError as exc:
        print(f"  Manchester refused the request ({exc}); using the figures read in a browser on 2 Sep 2026")
        return [{"school_name": n, "phase_hint": "secondary", "last_distance_miles": d} for n, d in _MANCHESTER_FALLBACK]
    text = re.sub(r"<(br|/p|/h[1-6]|/li|/div)[^>]*>", "\n", resp.text, flags=re.I)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"[ \t]+", " ", text)
    records = []
    for m in re.finditer(r"([^\n]+)\n\s*Places: *\d+\s*\n\s*Offers made:[^\n]*\n\s*Unfilled places:[^\n]*\n\s*([^\n]+)", text):
        name, rule = m.group(1).strip(), m.group(2)
        miles = re.search(r"live within ([0-9.]+) *miles", rule)
        if miles:
            records.append({"school_name": name, "phase_hint": "secondary", "last_distance_miles": float(miles.group(1))})
    return records


def _devon_breakdown(url: str, phase_hint: str) -> list[dict]:
    """Devon's allocation breakdowns: one row per school with the places,
    the offers by category, whether the school was oversubscribed and,
    when the distance tiebreaker was reached, "Distance (Metres) of Last
    Offered Place". Straight line, as the council measures. A blank
    distance means every applicant in the last category got in, so no
    figure is recorded and the row is skipped."""
    print(f"  Downloading {url}")
    resp = httpx.get(url, timeout=90, follow_redirects=True, headers=HEADERS)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)
    rows = list(wb.worksheets[0].iter_rows(values_only=True))
    header = [str(c or "").strip().lower() for c in rows[0]]
    dist_col = next(i for i, h in enumerate(header) if h.startswith("distance (metres)"))
    name_col = next(i for i, h in enumerate(header) if h.startswith("school"))
    records = []
    for row in rows[1:]:
        name = row[name_col] if len(row) > name_col else None
        dist = row[dist_col] if len(row) > dist_col else None
        if not isinstance(name, str) or dist in (None, "", 0):
            continue
        try:
            metres = float(dist)
        except (TypeError, ValueError):
            continue
        name = name.replace("\xa0", " ").replace("(see notes)", "").strip()
        records.append({"school_name": name, "phase_hint": phase_hint, "last_distance_miles": metres / _METRES_PER_MILE})
    return records


def fetch_devon_secondary() -> list[dict]:
    """Devon County Council, "Allocation Breakdown for each Secondary
    School" (Excel on the council's public SharePoint, linked from its
    apply-for-a-secondary-school-place page). The sheet's notes date it
    to National Offer Day 1 March 2024, so the September 2024 intake.
    Ten of the county's secondaries reached the distance tiebreaker."""
    return _devon_breakdown(
        "https://devoncc.sharepoint.com/:x:/s/PublicDocs/Education/IQBZyx4kiUt9TaSlpuswF_HMAeweeiSpomSujcuDfw4GCoU?e=fyTTxE&download=1",
        "secondary",
    )


def fetch_devon_primary() -> list[dict]:
    """Devon County Council, "2026 primary allocation breakdown" (Excel on
    the council's public SharePoint, linked from its apply-for-a-primary-
    school-place page): reception, September 2026 intake, 43 schools with
    a last-offered distance."""
    return _devon_breakdown(
        "https://devoncc.sharepoint.com/:x:/s/PublicDocs/Education/IQAZGIaqfE6QR5cX3ROIlHBfARZCcys1O1yNYsjKocw0CGo?e=uU8aHH&download=1",
        "reception",
    )


_NOTTS_PDFS = [
    ("https://www.nottinghamshire.gov.uk/media/uwzdg5rb/ashfield-primary.pdf", "primary"),
    ("https://www.nottinghamshire.gov.uk/media/mqwkbpre/bassetlaw-primary.pdf", "primary"),
    ("https://www.nottinghamshire.gov.uk/media/oxtfwjap/broxtowe-primary.pdf", "primary"),
    ("https://www.nottinghamshire.gov.uk/media/j2igrimw/gedling-primary.pdf", "primary"),
    ("https://www.nottinghamshire.gov.uk/media/05ncou5q/mansfield-primary.pdf", "primary"),
    ("https://www.nottinghamshire.gov.uk/media/tqlpiam5/newark-primary.pdf", "primary"),
    ("https://www.nottinghamshire.gov.uk/media/u2yfcd5y/rushcliffe-primary.pdf", "primary"),
    ("https://www.nottinghamshire.gov.uk/media/45wdtzga/secondary.pdf", "secondary"),
]
_NOTTS_AREAS = ("Ashfield", "Bassetlaw", "Broxtowe", "Gedling", "Mansfield", "Newark", "Rushcliffe")


def fetch_nottinghamshire() -> list[dict]:
    """Nottinghamshire County Council, "Schools information" PDFs, one per
    district for primary plus one for secondary, on the council's
    admissions pages (the media URLs rotate with each year's edition, so
    check the schools-information page before re-running). The last
    pages of each carry the "allocation summary" for the previous
    round: one row per school with preferences, places, offers by
    criterion and "distance of last preference offered (miles)", based
    on national offer day (16 April 2025 for reception and year 3,
    March 2025 for year 7). Nottinghamshire measures walking distance
    for most schools and a straight line for some, as each school's
    criteria say; the figure here is the one the council published.
    The council prints a distance for every school, full or not, and
    for a school with places to spare it is merely the furthest
    applicant (Sacred Heart in Mansfield shows 103 miles), so only rows
    where the places allocated reached the published admission number
    are kept. The text layout is a flattened table: the school name is
    everything between the district name and the first count; the first
    three counts are preferences, PAN and allocated; the distance is the
    row's only decimal number. The secondary PDF's summary table is not
    extractable as text, so Nottinghamshire is primary only.
    """
    records = []
    seen = set()
    row_re = re.compile(
        r"^(?:%s)\s+(.+?)\s+(\d+)\s+(\d+)\s+(\d+)(?:\s+\S+)*?\s+(\d+\.\d{1,3})(?:\s|$)" % "|".join(_NOTTS_AREAS)
    )
    for url, phase_hint in _NOTTS_PDFS:
        print(f"  Downloading {url}")
        resp = httpx.get(url, timeout=90, follow_redirects=True, headers=HEADERS)
        resp.raise_for_status()
        with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                if "allocation summary" not in text.lower() or "(miles)" not in text:
                    continue
                for line in text.splitlines():
                    m = row_re.match(line.strip())
                    if not m:
                        continue
                    name, pan, allocated, miles = m.group(1).strip(), int(m.group(3)), int(m.group(4)), float(m.group(5))
                    # A name that swallowed a count means the row wrapped;
                    # keep names that end in a word. A school that did not
                    # fill has no cut-off, whatever the column says.
                    # Two full primaries show 98 and 101 miles: the last offer
                    # went under a criterion that ignores distance, so the
                    # column is not a cut-off. No primary admits on distance
                    # from twenty miles.
                    if not re.search(r"[A-Za-z)]$", name) or miles <= 0 or allocated < pan or miles > 20:
                        continue
                    key = (name.lower(), phase_hint)
                    if key in seen:
                        continue
                    seen.add(key)
                    records.append({"school_name": name, "phase_hint": phase_hint, "last_distance_miles": miles})
    return records

_AUTHORITIES = [
    ("Nottinghamshire", "2025/26", fetch_nottinghamshire),
    ("Devon", "2024/25", fetch_devon_secondary),
    ("Devon", "2026/27", fetch_devon_primary),
    ("Manchester", "2026/27", fetch_manchester),
    ("Sheffield", "2025/26", fetch_sheffield),
    ("Essex", "2025/26", fetch_essex),
    ("Hounslow", "2025/26", fetch_hounslow),
    ("Wandsworth", "2024/25", fetch_wandsworth),
    ("Brent", "2023/24", fetch_brent),
    ("Bristol, City of", "varies", fetch_bristol),
    ("Leeds", "2024/25", fetch_leeds),
    ("Kirklees", "2024/25", fetch_kirklees),
    ("Ealing", "2024/25", fetch_ealing),
    ("Hackney", "2024/25", fetch_hackney),
    ("Solihull", "varies", fetch_solihull),
    ("Harrow", "2024/25", fetch_harrow),
    ("Gloucestershire", "2024/25", fetch_gloucestershire),
    ("Birmingham", "2023/24", fetch_birmingham),
    ("Newcastle upon Tyne", "varies", fetch_newcastle),
    ("Surrey", "2026/27", fetch_surrey),
    ("Hertfordshire", "varies", fetch_hertfordshire),
    ("Oxfordshire", "2026/27", fetch_oxfordshire),
    ("Buckinghamshire", "2026/27", fetch_buckinghamshire),
    ("Cambridgeshire", "2026/27", fetch_cambridgeshire),
    ("Wokingham", "2025/26", fetch_wokingham),
    ("Coventry", "varies", fetch_coventry),
    ("Milton Keynes", "2025/26", fetch_milton_keynes),
    ("Windsor and Maidenhead", "2025/26", fetch_windsor_and_maidenhead),
    ("Stockport", "2026/27", fetch_stockport),
    ("Portsmouth", "varies", fetch_portsmouth),
    ("Peterborough", "2026/27", fetch_peterborough),
    ("North Somerset", "2024/25", fetch_north_somerset),
    ("Somerset", "2025/26", fetch_somerset),
    ("Dorset", "2025/26", fetch_dorset),
    ("Worcestershire", "2026/27", fetch_worcestershire),
    ("Cheshire East", "2025/26", fetch_cheshire_east),
    ("Suffolk", "2026/27", fetch_suffolk),
    ("East Sussex", "varies", fetch_east_sussex),
    ("West Sussex", "2025/26", fetch_west_sussex),
    ("County Durham", "varies", fetch_durham),
    ("Derby", "2026/27", fetch_derby),
    ("Leicester", "varies", fetch_leicester),
    ("Sandwell", "varies", fetch_sandwell),
    ("Dudley", "2025", fetch_dudley),
    ("Richmond upon Thames", "2024/25", fetch_richmond_upon_thames),
    ("Kingston upon Thames", "2024/25", fetch_kingston_upon_thames),
    ("Southwark", "2025/26", fetch_southwark),
    ("Lambeth", "2024/25", fetch_lambeth),
    ("Waltham Forest", "varies", fetch_waltham_forest),
    ("Lewisham", "2025/26", fetch_lewisham),
    ("Merton", "varies", fetch_merton),
    ("Greenwich", "2024/25", fetch_greenwich),
    ("Sutton", "2025/26", fetch_sutton),
    ("Haringey", "varies", fetch_haringey),
    ("Calderdale", "2026/27", fetch_calderdale),
    ("Walsall", "2025/26", fetch_walsall),
    ("Oldham", "2025/26", fetch_oldham),
    ("Wigan", "2025/26", fetch_wigan),
    ("Sefton", "2025/26", fetch_sefton),
    ("Newham", "2025/26", fetch_newham),
    ("Bexley", "varies", fetch_bexley),
    ("Southampton", "2024/25", fetch_southampton),
    ("Islington", "2026/27", fetch_islington),
    ("Havering", "2025/26", fetch_havering),
    ("Hillingdon", "2025/26", fetch_hillingdon),
    ("Middlesbrough", "2025/26", fetch_middlesbrough),
    ("Hartlepool", "2025/26", fetch_hartlepool),
    ("Southend-on-Sea", "2025/26", fetch_southend),
    ("Bracknell Forest", "2025/26", fetch_bracknell_forest),
    ("Tameside", "varies", fetch_tameside),
    ("Gloucestershire", "2025/26", fetch_gloucestershire),
    ("Warwickshire", "2025/26", fetch_warwickshire),
    ("Staffordshire", "varies", fetch_staffordshire),
    ("Kirklees", "2025/26", fetch_kirklees),
    ("Cheshire West and Chester", "2025/26", fetch_cheshire_west_and_chester),
    ("Bristol, City of", "varies", fetch_bristol),
    ("North Yorkshire", "2025/26", fetch_north_yorkshire),
    ("Reading", "2025/26", fetch_reading),
    ("Bury", "varies", fetch_bury),
    ("Bolton", "varies", fetch_bolton),
    ("Salford", "2025/26", fetch_salford),
    ("Knowsley", "2026/27", fetch_knowsley),
    ("Brighton and Hove", "2025/26", fetch_brighton_and_hove),
    ("Bromley", "2025/26", fetch_bromley),
    ("Camden", "2024", fetch_camden),
    ("Westminster", "2024/25", fetch_westminster),
    ("Kensington and Chelsea", "2025/26", fetch_kensington_and_chelsea),
    ("Tower Hamlets", "2024/25", fetch_tower_hamlets),
    ("Wirral", "2025/26", fetch_wirral),
    ("Bedford", "2026/27", fetch_bedford),
    ("Central Bedfordshire", "2026/27", fetch_central_bedfordshire),
    ("West Northamptonshire", "2026/27", fetch_west_northamptonshire),
]


def _match_urn(school_name: str, candidates: dict[str, int]) -> int | None:
    normalized = _normalize_school_name(school_name)
    matches = difflib.get_close_matches(normalized, candidates.keys(), n=1, cutoff=_MATCH_CUTOFF)
    if matches:
        return candidates[matches[0]]
    # Councils shorten names ("Meadowhead" for "Meadowhead School Academy
    # Trust"). A whole-word prefix that fits exactly one school in the
    # authority is safe; two or more and the row is left unmatched.
    prefixed = [urn for name, urn in candidates.items() if name.startswith(normalized + " ")]
    return prefixed[0] if len(prefixed) == 1 else None


def build_records(session) -> list[dict]:
    records = []
    for authority, academic_year, fetch_fn in _AUTHORITIES:
        print(f"Fetching {authority}...")
        try:
            rows = fetch_fn()
        except Exception as exc:
            # A single authority's source URL rotting/going down (these are
            # re-published every year at a new address, sometimes behind a
            # WAF that blocks this environment) shouldn't take down the
            # whole import - skip it and keep going, rather than lose every
            # other authority's data too.
            print(f"  SKIPPED - {authority} fetch failed: {exc}")
            continue
        print(f"  {len(rows)} schools with a distance figure in the source file")

        schools_in_la = session.execute(
            select(School.urn, School.name, School.phase)
            .join(SchoolDetail, SchoolDetail.urn == School.urn)
            .where(SchoolDetail.local_authority == authority)
        ).all()
        candidates = {_normalize_school_name(name): urn for urn, name, _ in schools_in_la}
        valid_urns = {urn for urn, _, _ in schools_in_la}
        # A fetcher that knows which intake a table describes can say so
        # ("phase_hint": "secondary", "primary" or "reception") and is
        # then matched only against schools of that phase, so a council's
        # short "Ecclesfield" finds the primary in the reception table and
        # the secondary in the Year 7 one. Reception also drops junior
        # schools, which have no reception year.
        _PHASES = {
            "secondary": {"Secondary", "Middle deemed secondary", "All-through"},
            "primary": {"Primary", "Middle deemed primary", "All-through"},
            "reception": {"Primary", "Middle deemed primary", "All-through"},
        }
        candidates_by_phase = {}
        for hint, phases in _PHASES.items():
            pool = {}
            for urn, name, phase in schools_in_la:
                if phase not in phases:
                    continue
                norm = _normalize_school_name(name)
                if hint == "reception" and re.search(r"\bjunior\b", norm) and not re.search(r"\b(infant|primary)\b", norm):
                    continue
                pool[norm] = urn
            candidates_by_phase[hint] = pool

        matched = 0
        unmatched = []
        for row in rows:
            # A source that publishes its own URN (e.g. Greenwich) is matched directly rather than
            # by fuzzy name, but still checked against this authority's real URNs rather than trusted blindly.
            # A distance of nothing is a council's way of writing "no place
            # was decided on distance" (Gloucestershire's 0.0) or a unit slip
            # (Sutton's 0.002 miles, four metres). Either way it is not a
            # figure, and shown as one it would top the national ranking.
            if (row.get("last_distance_miles") or 0) < 0.01:
                unmatched.append(f"{row['school_name']} (no usable distance: {row.get('last_distance_miles')})")
                continue
            if row.get("urn") in valid_urns:
                urn = row["urn"]
            else:
                pool = candidates_by_phase.get(row.get("phase_hint"), candidates)
                urn = _match_urn(row["school_name"], pool)
            if urn is None:
                unmatched.append(row["school_name"])
                continue
            records.append({
                "urn": urn,
                "academic_year": academic_year,
                "last_distance_miles": row["last_distance_miles"],
                "source_authority": authority,
            })
            matched += 1

        print(f"  matched {matched}/{len(rows)} to a school in our database (before de-duplication)")
        if unmatched:
            print(f"  unmatched, skipped rather than guessed: {unmatched}")

    # A school can appear twice within one authority's own source file
    # (e.g. split across a page boundary in the raw PDF text) - keep
    # the first occurrence per URN rather than let a duplicate insert
    # crash the whole import.
    by_urn: dict[int, dict] = {}
    duplicates = 0
    for r in records:
        if r["urn"] in by_urn:
            duplicates += 1
            continue
        by_urn[r["urn"]] = r
    if duplicates:
        print(f"Dropped {duplicates} duplicate URN(s) (same school matched more than once)")

    return list(by_urn.values())


def load_into_db(records: list[dict], only: list[str] | None = None) -> None:
    engine = _get_engine()
    Base.metadata.create_all(engine, tables=[SchoolAdmissionRadius.__table__])
    SessionLocal = sessionmaker(bind=engine)

    with SessionLocal() as session:
        if only:
            # Adding one council must not re-fetch 83 others, any of
            # whose URLs may have rotted since, and lose their rows.
            print(f"Replacing rows for {', '.join(only)} only...")
            session.query(SchoolAdmissionRadius).filter(
                SchoolAdmissionRadius.source_authority.in_(only)
            ).delete(synchronize_session=False)
        else:
            print("Clearing existing school_admission_radii table...")
            session.query(SchoolAdmissionRadius).delete()
        session.commit()

        print(f"Inserting {len(records)} rows...")
        if records:
            session.execute(SchoolAdmissionRadius.__table__.insert(), records)
            session.commit()


def main():
    """Usage: import_admission_radii.py [--only "Essex,Lincolnshire"]

    With --only, just those authorities are fetched and their rows
    replaced; everything else in the table is left exactly as it was."""
    global _AUTHORITIES
    only = None
    if "--only" in sys.argv:
        only = [a.strip() for a in sys.argv[sys.argv.index("--only") + 1].split(",") if a.strip()]
        _AUTHORITIES = [t for t in _AUTHORITIES if t[0] in only]
        missing = set(only) - {t[0] for t in _AUTHORITIES}
        if missing:
            sys.exit(f"not in the registry: {sorted(missing)}")
    engine = _get_engine()
    SessionLocal = sessionmaker(bind=engine)
    with SessionLocal() as session:
        records = build_records(session)
    load_into_db(records, only=only)
    print("Done.")


if __name__ == "__main__":
    main()
